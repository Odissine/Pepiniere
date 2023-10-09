import urllib.parse
import socket
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Q, Count, F, Value, Sum
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse, FileResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import get_template
from django.utils.html import format_html
from django_xhtml2pdf.utils import generate_pdf
from django.utils import timezone
from django.db.models.expressions import RawSQL

from datetime import datetime, date
from openpyxl import load_workbook
import pandas
import io
import locale
import xlsxwriter
from tablib import Dataset

from .utils import render_to_pdf
from .forms import *
from .models import *
from .core import *
from .resources import *
from onlineshop.models import *
from onlineshop.core import *
from cart.forms import CartAddProduitForm, CartUpdateForm, RemiseUpdateForm

locale.setlocale(locale.LC_ALL, 'fr_FR')


# AFFICHE LA LISTE DES COMMANDES ____________________________________________________________________________
@login_required
# @staff_member_required(login_url='account:error')
def order_list(request):

    if request.user.is_staff is False:
        orders = Commande.objects.filter(client__user=request.user)
    else:
        orders = Commande.objects.all()

    if 'o' in request.GET:
        order_value = request.GET['o']
        if order_value == 'total':
            orders = orders.annotate(total_order=Sum(F('Cartdbs__qte') * F('Cartdbs__prix'))).order_by('-total_order')
        elif order_value == 'qte':
            orders = orders.annotate(qte_order=Count('Cartdbs')).order_by('-qte_order')
        else:
            orders = orders.order_by('-'+str(order_value))

        request.session['o'] = request.GET['o']
    elif 'o' in request.session:
        order_value = request.session['o']
        if order_value == 'total':
            orders = orders.annotate(total_order=Sum(F('Cartdbs__qte') * F('Cartdbs__prix'))).order_by('-total_order')
        elif order_value == 'qte':
            orders = orders.annotate(qte_order=Count('Cartdbs')).order_by('-qte_order')
        else:
            orders = orders.order_by('-'+str(order_value))
    else:
        orders = orders.order_by('-date', 'statut')

    formAction = "order:order-list"
    form = SearchOrderForm()

    if request.method == 'GET':
        form = SearchOrderForm(request.GET)
        if form.is_valid():
            statut = form.cleaned_data['statut']
            clients = form.cleaned_data['clients']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            produits = form.cleaned_data['produits']
            especes = form.cleaned_data['especes']
            varietes = form.cleaned_data['varietes']
            portegreffes = form.cleaned_data['portegreffes']
            frais = form.cleaned_data['frais']
            inventaire = form.cleaned_data['inventaire']

            if statut.exists():
                orders = orders.filter(statut__in=statut)
            if clients.exists():
                orders = orders.filter(client__in=clients)
            if start_date is not None and start_date != "":
                orders = orders.filter(date__gte=start_date)
            if end_date is not None and end_date != "":
                orders = orders.filter(date__lte=end_date)
            if produits.exists():
                orders = orders.filter(Cartdbs__produit__in=produits)
            if especes.exists():
                orders = orders.filter(Cartdbs__produit__espece__in=especes)
            if varietes.exists():
                orders = orders.filter(Cartdbs__produit__variete__in=varietes)
            if portegreffes.exists():
                orders = orders.filter(Cartdbs__produit__portegreffe__in=portegreffes)
            if frais.exists():
                orders = orders.filter(frais__in=frais)
            if inventaire.exists():
                orders = orders.filter(inventaire__in=inventaire)

        if 'max_val' in request.GET:
            try:
                max_value = int(request.GET['max_val'])
            except:
                max_value = 5
            orders = get_orders_items_max(max_value)

    paginator = Paginator(orders, 50)
    get_data = request.GET.copy()
    page = get_data.pop('page', None)
    o = get_data.pop('o', None)
    if 'page' in request.GET:
        page = request.GET['page']
    elif 'page' in request.session:
        page = request.session['page']

    try:
        orders_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        orders_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        orders_page = paginator.page(paginator.num_pages)

    context = {'orders': orders_page,
               'orders_list': orders,
               'paginate': True,
               'form': form,
               'formAction': formAction,
               'query_string': get_data.urlencode(),
               }
    return render(request, 'order/list.html', context)


# AFFICHE LE DETAIL D'UNE COMMANDE ____________________________________________________________________________
@login_required
def order_detail(request, id):
    commande = get_object_or_404(Commande, id=id)
    if not request.user.is_staff:
        commande_users = Commande.objects.filter(client__user=request.user).all()
        if not commande in commande_users:
            messages.error(request, "Commande inexistante !")
            return redirect('order:order-list')

    produits = Cartdb.objects.filter(commande=commande)

    frais = Frais.objects.all()
    tvas = Tva.objects.filter(active=True)
    form = FormAddProduit(request.POST or None, order=commande)

    context = {
        'commande': commande,
        'produits': produits,
        'frais': frais,
        'tvas': tvas,
        'form': form,
    }

    return render(request, 'order/detail.html', context)


# PASSAGE D'UNE COMMANDE EN ATTENTE EN COMMANDE EN COURS _______________________________________________________________
@login_required
@staff_member_required
def order_accept(request, id):
    previous = request.META.get('HTTP_REFERER').split("/")[-1]
    try:
        order = get_object_or_404(Commande, pk=id)
    except:
        messages.error(request, "Commande inexistante !")
        return redirect('order:order-detail', id)

    if order.statut.nom != "En attente":
        messages.error(request, "Action impossible avec cette commande (problème de statut) !")
        return redirect('order:order-detail', id)

    items = Cartdb.objects.filter(commande=order)
    statut = Statut.objects.get(nom='En cours')
    nb_produit = 0
    message = ""
    for item in items:
        stock_virtuel = item.produit.stock_bis
        if stock_virtuel - item.qte < 0:
            nb_produit += 1
            message = message + "<li>" + item.produit.nom + "</li>"

    if nb_produit > 0:
        message = format_html("Stock insuffisant pour accepter la commande pour les produits suivants : <ul>" + message + "</ul>")
        messages.error(request, message)
        if previous == "manage":
            return redirect('order:manage-order')
        else:
            return redirect('order:order-detail', id)

    for item in items:
        produit = item.produit
        old_stock = produit.stock_bis
        new_stock = produit.stock_bis - item.qte
        produit.stock_bis = produit.stock_bis - item.qte

        # LOG EDIT PRODUIT (STOCK)
        log_produit("Produit", str(request.user), produit.pk, order.pk, 'Update', 'sb', old_stock, new_stock)
        produit.save()

    old_statut = {order.statut.nom}
    new_statut = statut.nom
    Commande.objects.filter(pk=id).update(statut=statut, date_update=datetime.now())
    log_order("Commande", str(request.user), order.pk, 'Update', 'statut', old_statut, new_statut)

    message = "Commande en attente acceptée avec succès à partir de celle réalisée par %s" % order.client
    messages.success(request, message)
    if previous == "manage":
        return redirect('order:manage-order')
    else:
        return redirect('order:order-detail', id)


# STATUT VALIDEE D'UNE COMMANDE EN COURS ___________________________________________________________________________
@login_required
@staff_member_required
def order_valid(request, id):
    try:
        order = get_object_or_404(Commande, pk=id)
    except:
        messages.error(request, "Commande inexistante !")
        return redirect('order:order-detail', id)

    if order.statut.nom != "En cours":
        messages.error(request, "Action impossible avec cette commande (problème de statut) !")
        return redirect('order:order-detail', id)

    statut = Statut.objects.get(nom='Validée')
    old_value = order.statut.nom
    new_value = statut.nom
    Commande.objects.filter(pk=id).update(statut=statut, date_update=datetime.now())
    log_order("Commande", str(request.user), order.pk, 'Update', 'statut', old_value, new_value)

    message = "Commande validée avec succès :)"
    messages.success(request, message)
    return redirect('order:order-detail', id)


# PASSAGE D'UNE PRE COMMANDE EN COMMANDE EN COURS _______________________________________________________________________________
@login_required
@staff_member_required
def order_pre_valid(request, id):
    try:
        order = get_object_or_404(Commande, pk=id)
    except:
        messages.error(request, "Commande inexistante !")
        return redirect('order:order-detail', id)

    if order.statut.nom != "Pré-commande":
        messages.error(request, "Action impossible avec cette commande (problème de statut) !")
        return redirect('order:order-detail', id)

    items = Cartdb.objects.filter(commande=order)
    statut = Statut.objects.get(nom='En cours')
    nb_produit = 0
    message = ""
    for item in items:
        stock_virtuel = item.produit.stock_bis
        if stock_virtuel - item.qte < 0:
            nb_produit += 1
            message = message + "<li>" + item.produit.nom + "</li>"

    if nb_produit > 0:
        message = format_html("Stock insuffisant pour permettre la création de la commande pour les produits suivants : <ul>" + message + "</ul>")
        messages.error(request, message)
        return redirect('order:order-detail', id)

    for item in items:
        produit = item.produit
        old_bis = produit.stock_bis
        old_future = produit.stock_future
        new_bis = produit.stock_bis - item.qte
        new_future = produit.stock_bis - item.qte
        produit.stock_bis = produit.stock_bis - item.qte
        produit.stock_future = produit.stock_future - item.qte
        produit.save()

        # LOG EDIT PRODUIT (STOCK)
        log_produit("Produit", str(request.user), produit.pk, order.pk, 'Update', 'sb', old_bis, new_bis)
        log_produit("Produit", str(request.user), produit.pk, order.pk, 'Update', 'sp', old_future, new_future)

    inventaire = Inventaire.objects.get(start_date__lte=datetime.now(), end_date__gte=datetime.now())
    old_inventaire = order.inventaire
    old_statut = order.statut
    Commande.objects.filter(pk=id).update(statut=statut, date_update=datetime.now(), inventaire=inventaire)
    # LOG EDIT ORDER
    log_order("Commande", str(request.user), order.pk, 'Update', 'statut', old_statut, statut)
    log_order("Commande", str(request.user), order.pk, 'Update', 'inventaire', old_inventaire, inventaire)

    message = "Commande créée avec succès à partir de la Pré-commande !"
    messages.success(request, message)
    return redirect('order:order-detail', id)


# PASSAGE DE TOUTES LES PRE-COMMANDES EN COMMANDES EN COURS ________________________________________________________________
@login_required
@staff_member_required
def all_order_pre_valid(request, action=None):
    orders = Commande.objects.filter(statut__nom="Pré-commande")

    if action == 'True':
        statut = Statut.objects.get(nom='En cours')
        nb_produit = 0
        message = ""
        for order in orders:
            items = Cartdb.objects.filter(commande=order)

            for item in items:
                stock_virtuel = item.produit.stock_bis
                if stock_virtuel - item.qte < 0:
                    nb_produit += 1
                    message += "<li>" + item.produit.nom + "</li>"

        if nb_produit > 0:
            message = format_html("Stock insuffisant pour permettre la création de la commande pour les produits suivants : <ul>" + message + "</ul>")
            messages.error(request, message)
            return redirect('order:order-administration')
        try:
            inventaire = Inventaire.objects.get(actif=True)
        except:
            inventaire = Inventaire.objects.filter(start_date__lte=datetime.now(), end_date__gte=datetime.now()).last()
        for order in orders:
            items = Cartdb.objects.filter(commande=order)
            for item in items:
                produit = item.produit
                old_bis = produit.stock_bis
                new_bis = produit.stock_bis - item.qte
                old_future = produit.stock_future
                new_future = produit.stock_future - item.qte
                produit.stock_bis = produit.stock_bis - item.qte
                produit.stock_future = produit.stock_future - item.qte
                produit.save()
                log_produit("Produit", str(request.user), produit.pk, order.pk, 'Update', 'sb', old_bis, new_bis)
                log_produit("Produit", str(request.user), produit.pk, order.pk, 'Update', 'sp', old_future, new_future)

            old_inventaire = order.inventaire.start_date.strftime('%Y') + '-' + order.inventaire.end_date.strftime('%Y')
            old_statut = order.statut.nom
            order.statut = statut
            order.date_update = datetime.now()
            # order.date = datetime.now()
            order.inventaire = inventaire
            order.save()
            log_order("Commande", str(request.user), order.pk, 'Update', 'inventaire', old_inventaire, inventaire.start_date.strftime('%Y') + '-' + inventaire.end_date.strftime('%Y'))
            log_order("Commande", str(request.user), order.pk, 'Update', 'statut', old_statut, statut.nom)

        message = "Commandes créées avec succès à partir des Pré-commandes !"
        messages.success(request, message)
        return redirect('order:order-administration')

    context = {
        'previous-page': 'order:order-administration',
        'orders': orders,
    }
    return render(request, 'order/form_pre_order.html', context)


# CREATION D'UNE PRE-COMMANDE _______________________________________________________________________________________________
@login_required
@staff_member_required
def pre_order_create(request, id):
    try:
        order = get_object_or_404(Commande, pk=id)
    except:
        messages.error(request, "Commande inexistante !")
        return redirect('order:order-detail', id)

    items = Cartdb.objects.filter(commande=order)
    statut = Statut.objects.get(nom='Pré-commande')
    order.pk = None
    order.id = None
    order.save()
    order.date_update = timezone.now()
    order.date = datetime.now()
    order.statut = statut
    order.save()
    inventaire = set_inventaire_for_pre_order(order.id)

    # LOG CREATION DE LA PRE-COMMANDE
    log_order("Commande", str(request.user), order.pk, 'Create', 'statut', '', statut.nom)
    log_order("Commande", str(request.user), order.pk, 'Create', 'client', '', str(order.client.nom) + ' ' + str(order.client.prenom))
    log_order("Commande", str(request.user), order.pk, 'Create', 'tva', 0, float(order.tva.tva))
    log_order("Commande", str(request.user), order.pk, 'Create', 'inventaire', '', inventaire.start_date.strftime('%Y')+'-'+inventaire.end_date.strftime('%Y'),)
    if order.frais:
        log_order("Commande", str(request.user), order.pk, 'Create', 'frais', '', order.frais.nom)
        log_order("Commande", str(request.user), order.pk, 'Create', 'montant_frais', 0, float(order.montant_frais))

    for item in items:
        produit = Produit.objects.get(id=item.produit.id)
        item.id = None
        item.pk = None
        item.commande = order
        item.save()
        old_future = produit.stock_future
        new_future = produit.stock_future + item.qte
        produit.stock_future += item.qte
        produit.save()

        # LOG CREATION DE LA COMMANDE (AJOUT PRODUIT)
        log_cart("Cart", str(request.user), item.pk, order.pk, produit.pk, 'Create', 'qte', 0, item.qte)
        log_cart("Cart", str(request.user), item.pk, order.pk, produit.pk, 'Create', 'prix', '', float(item.prix))

        # LOG CREATION DE LA COMMANDE (MAJ STOCK)
        print(old_future, new_future)
        log_produit("Produit", str(request.user), produit.pk, order.pk, 'Create', 'sp', old_future, new_future)

    message = "Pré-Commande créée avec succès à partir de la Commande selectionnée !"
    messages.success(request, message)

    previous = request.META.get('HTTP_REFERER').split("/")[-1]
    if previous == "manage":
        return redirect('order:manage-order')
    else:
        return redirect('order:order-detail', order.id)


# MISE A JOUR D'UNE COMMANDE (QTE et PRIX) _______________________________________________________________________________________________
@login_required
@staff_member_required
def order_update_qte_prix(request, id):
    produit_commande = get_object_or_404(Cartdb, id=id)
    try:
        admin_mode = AccessMode.objects.get(user=request.user).admin
    except:
        admin_mode = False

    if request.method == "POST":
        try:
            prix = float(request.POST['prix'].replace(',', '.'))
            qte = int(request.POST['qte'])
        except Exception as e:
            message = format_html("Une erreur s'est produite : <br>" + str(e))
            messages.error(request, message)
            return redirect('order:order-detail', produit_commande.commande.id)

        if isinstance(prix, float) and isinstance(qte, int):
            # MISE A JOUR DES STOCK VIRTUELS
            produit = Produit.objects.get(id=produit_commande.produit.id)

            # LOG MODIFICATION COMMANDE (EDITION PRODUIT PRIX / QTE)
            old_qte = produit_commande.qte
            old_prix = produit_commande.prix

            produit_commande.qte = qte
            produit_commande.prix = prix
            produit_commande.save()

            if old_qte != qte:
                log_cart("Cart", str(request.user), produit_commande.pk, produit_commande.commande.pk, produit_commande.produit.pk, 'Update', 'qte', old_qte, qte)
            if old_prix != prix:
                log_cart("Cart", str(request.user), produit_commande.pk, produit_commande.commande.pk, produit_commande.produit.pk, 'Update', 'prix', float(old_prix), float(prix))
            # PRE-COMMANDE
            if produit_commande.commande.statut.nom == "Pré-commande":
                # ON MET A JOUR LE PRODUIT DE LA COMMANDE AVEC LA NOUVELLE QUANTITE ET LE NOUVEAU PRIX
                old_stock = produit.stock_future
                new_stock = produit.stock_future - old_qte + qte
                produit.stock_future = new_stock
                produit.save()
                # LOG STOCK UPDATE
                if old_stock != new_stock:
                    log_produit("Produit", str(request.user), produit_commande.produit.pk, produit_commande.commande.pk, 'Update', 'sp', old_stock, new_stock)

                # ON MET A JOUR LA COMMANDE AVEC LE NOUVEAU TOTAL
                Commande.objects.filter(pk=produit_commande.commande.id).update(date_update=datetime.now())

                message = format_html("Mise à jour quantités / prix effectuée avec succès !")
                messages.success(request, message)

            # EN COURS / VALIDEE
            else:
                old_stock = produit.stock_bis
                new_stock = produit.stock_bis + old_qte - qte

                if new_stock >= 0:
                    produit.stock_bis = new_stock
                    produit.save()
                    if old_stock != new_stock:
                        log_produit("Produit", str(request.user), produit.pk, produit_commande.commande.pk, 'Update', 'sb', old_stock, new_stock)

                    # ON MET A JOUR LA COMMANDE AVEC LE NOUVEAU TOTAL
                    Commande.objects.filter(pk=produit_commande.commande.id).update(date_update=datetime.now())

                    message = format_html("Mise à jour quantités / prix effectuée avec succès !")
                    messages.success(request, message)
                else:
                    message = format_html("Stock insuffisant !<br/> Modification impossible avec cette quantité !")
                    messages.error(request, message)
                    return redirect('order:order-detail', produit_commande.commande.id)
        else:
            message = "Une erreur s'est produite !"
            messages.error(request, message)

    return redirect('order:order-detail', produit_commande.commande.id)


# MISE A JOUR DE LA REMISE SUR UNE COMMANDE __________________________________________________________________
@login_required
@staff_member_required
def order_update_remise(request, id):
    try:
        remise = locale.atof(request.POST['remise'])
        order = Commande.objects.get(pk=id)
    except Exception as e:
        message = format_html("Une erreur s'est produite : <br>" + str(e))
        messages.error(request, message)
        return redirect('order:order-detail', id)
    old_remise = order.remise
    new_remise = remise
    order.update(remise=remise)
    log_order("Commande", str(request.user), order.pk, 'Edit', 'remise', float(old_remise), float(new_remise))

    message = "Remise modifiée avec succès !"
    messages.success(request, message)
    return redirect('order:order-detail', id)


# MISE A JOUR DES FRAIS SUR UNE COMMANDE (AJAX - CHANGE INPUT)
@login_required
@staff_member_required
def order_update_frais(request, id):
    if request.method == "POST":
        try:
            commande = get_object_or_404(Commande, pk=id)
        except:
            messages.error(request, "Commande inexistante !")
            return redirect('order:order-detail', id)

        frais_montant = request.POST["fraisMontant"]
        frais_type = request.POST["frais_type"]

        if not frais_montant:
            frais_montant = 0
        else:
            try:
                frais_montant = locale.atof(frais_montant)
                try:
                    frais_type = Frais.objects.get(pk=frais_type)
                except Exception as eFrais:
                    frais_type = None
                    frais_montant = 0

            except Exception as e:
                message = format_html("Une erreur s'est produite : <br>" + str(e))
                messages.error(request, message)
                return redirect('order:order-detail', id)
        old_frais = ""
        old_montant_frais = ""
        if commande.frais:
            old_frais = commande.frais.nom
            old_montant_frais = commande.montant_frais
        commande.frais = frais_type
        commande.date_update = datetime.now()
        commande.montant_frais = frais_montant
        commande.save()

        log_order("Commande", str(request.user), commande.pk, 'Edit', 'frais', old_frais, frais_type.nom)
        log_order("Commande", str(request.user), commande.pk, 'Edit', 'montant_frais', float(old_montant_frais), float(frais_montant))

        message = "Frais modifiés avec succès !"
        messages.success(request, message)

    return redirect('order:order-detail', id)


# SUPPRESSION D'UN PRODUIT DE LA COMMANDE ____________________________________________________________________________________
@login_required
@staff_member_required
def order_product_remove(request, id):
    item = get_object_or_404(Cartdb, id=id)

    commande = get_object_or_404(Commande, pk=item.commande.id)
    old_future = item.produit.stock_future
    old_final = item.produit.stock
    old_bis = item.produit.stock_bis
    if commande.statut.nom == "En cours" or commande.statut.nom == "Validée":
        stock = item.produit.stock_bis + item.qte
        Produit.objects.filter(pk=item.produit.id).update(stock_bis=stock)
        log_produit("Produit", str(request.user), item.produit.pk, commande.pk, 'Update', 'sb', old_bis, stock)

    if commande.statut.nom == "Pré-commande":
        stock = item.produit.stock_future - item.qte
        Produit.objects.filter(pk=item.produit.id).update(stock_future=stock)
        log_produit("Produit", str(request.user), item.produit.pk, commande.pk, 'Update', 'sp', old_future, stock)

    old_value_cart = item.qte
    new_value_cart = 0
    log_cart("Cart", str(request.user), item.pk, commande.pk, item.produit.pk, 'Delete', 'qte', old_value_cart, new_value_cart)
    item.delete()

    message = "Suppression du produit effectuée avec succès !"
    messages.success(request, message)

    return redirect('order:order-detail', item.commande.id)


# ANNULATION D'UNE COMMANDE AVEC MISE A JOUR DE LA DATE ET DU STATUT ET DES STOCKS SI NECESSAIRE
@login_required
def order_cancel(request, id):
    order = get_object_or_404(Commande, pk=id)
    statut = Statut.objects.get(nom='Annulée')
    commande_statut = order.statut
    old_value_order = order.statut.nom
    new_value_order = statut.nom

    Commande.objects.filter(pk=id).update(statut=statut, date_update=datetime.now())
    log_order("Commande", str(request.user), order.pk, 'Cancel', 'statut', old_value_order, new_value_order)

    items = Cartdb.objects.filter(commande=order)
    message = ""
    for item in items:
        produit = Produit.objects.get(pk=item.produit.id)
        old_final = produit.stock
        old_bis = produit.stock_bis
        old_future = produit.stock_future
        if commande_statut.nom == "En attente":
            message = "Commande en attente annulée avec succès !"
        elif commande_statut.nom == "Pré-commande":
            old_qte = produit.stock_future - item.qte
            Produit.objects.filter(pk=item.produit.id).update(stock_future=old_qte)
            message = "Pré-commande annulée avec succès !"
            log_produit("Produit", str(request.user), produit.pk, order.pk, 'Cancel', 'sp', old_future, old_qte)
        elif commande_statut.nom == "Validée" or commande_statut.nom == "En cours":
            old_qte = produit.stock_bis + item.qte
            Produit.objects.filter(pk=item.produit.id).update(stock_bis=old_qte)
            message = "Commande annulée avec succès !"
            log_produit("Produit", str(request.user), produit.pk, order.pk, 'Cancel', 'sb', old_bis, old_qte)
        else:
            old_qte = produit.stock + item.qte
            Produit.objects.filter(pk=item.produit.id).update(stock=old_qte)
            message = "Commande annulée avec succès !"
            log_produit("Produit", str(request.user), produit.pk, order.pk, 'Cancel', 'sf', old_final, old_qte)
    messages.success(request, message)
    return redirect('order:order-detail', id)


# COMMANDE TERMINEE AVEC MISE A JOUR DE LA DATE ET DU STATUT
@login_required
@staff_member_required
def order_end(request, id):
    order = get_object_or_404(Commande, pk=id)
    statut = Statut.objects.get(nom='Terminée')

    if order.statut.nom != "Validée" and order.statut.nom != "En cours":
        message = format_html("Impossible de passer la commande en statut \"Terminée\"<br>Elle doit d'abord passer par le statut validée ou en cours !")
        messages.error(request, message)
        return redirect('order:order-detail', id)

    items = Cartdb.objects.filter(commande=order)
    for item in items:
        produit = Produit.objects.get(pk=item.produit.id)
        new_qte = produit.stock - item.qte
        if new_qte < 0:
            message = format_html("Impossible de passer la commande en statut \"Terminée\"<br>Stock Final insuffisant !")
            messages.error(request, message)
            return redirect('order:order-detail', id)

    old_value_order = order.statut.nom
    new_value_order = statut.nom
    log_order("Commande", str(request.user), order.pk, 'End', 'statut', old_value_order, new_value_order)

    Commande.objects.filter(pk=id).update(statut=statut, date_update=datetime.now())
    for item in items:
        old_qte = item.produit.stock
        produit = Produit.objects.get(pk=item.produit.id)
        new_qte = produit.stock - item.qte
        log_produit("Produit", str(request.user), produit.pk, order.pk, 'End', 'sf', old_qte, new_qte)

        Produit.objects.filter(pk=item.produit.id).update(stock=new_qte)

    message = "Commande terminée avec succès !"
    messages.success(request, message)
    return redirect('order:order-detail', id)


@login_required
def order_print(request, id, *args, **kwargs):
    try:
        mode = request.GET['mode']
    except:
        mode = "1"
    path_pdf = 'pdf/facture.html'

    if mode == "1":
        type_pdf = "Facture"
    elif mode == "2":
        type_pdf = "Commande"
    else:
        type_pdf = "Devis"

    template = get_template(path_pdf)
    commande = get_object_or_404(Commande, id=id)
    items = Cartdb.objects.filter(commande=commande)
    image = settings.STATIC_ROOT + '/img/logo_facture.png'

    context = {
        'type': type_pdf,
        'commande': commande,
        'items': items,
        'image': image,
    }
    # return render(request, path_pdf, context)

    pdf = render_to_pdf(path_pdf, context_dict=context)

    if pdf:
        response = HttpResponse(content_type='application/pdf')
        result = generate_pdf(path_pdf, file_object=response, context=context)

        filename = type_pdf + "_%s_%s.pdf" % (commande.id, commande.date.strftime('%Y'))
        content = "inline; filename=%s" % filename
        download = request.GET.get("download")
        if download == 1:
            content = "attachment; filename=%s" % filename
        response['Content-Disposition'] = content
        return response
        # return result
    return HttpResponse("Not found")


# Modal Search for Client by Name/Firstname for Add in Order
@login_required
@staff_member_required
def order_search_client(request):
    nom = request.POST.get("recipient-nom")
    prenom = request.POST.get("recipient-prenom")

    clients = Client.objects.all().values()

    if nom == "" and prenom == "":
        clients = Client.objects.all().values()
    if nom == "":
        clients = Client.objects.filter(prenom__icontains=prenom).values()
    if prenom == "":
        clients = Client.objects.filter(nom__icontains=nom).values()

    return JsonResponse({"clients_json": list(clients)})


# Modal Search for Commande by ID for Add Product in
@login_required
@staff_member_required
def order_search_order(request):
    order_id = request.POST.get("recipient-order")
    orders = Commande.objects.all().values()

    if order_id != "":
        orders = Commande.objects.filter(pk=order_id)
    # print(orders)

    return JsonResponse({"orders_json": orders})


@login_required
@staff_member_required
def order_etiquettes(request):
    formAction = 'order:order-etiquettes'
    form = SearchOrderForm(request.GET or None)

    if 'o' in request.GET:
        order_value = request.GET['o']
        if order_value == 'total':
            orders = Commande.objects.all().annotate(total_order=Sum(F('Cartdbs__qte') * F('Cartdbs__prix'))).order_by('-total_order')
        elif order_value == 'qte':
            orders = Commande.objects.all().annotate(qte_order=Count('Cartdbs')).order_by('-qte_order')
        else:
            orders = Commande.objects.all().order_by(order_value)

        request.session['o'] = request.GET['o']
    elif 'o' in request.session:
        order_value = request.session['o']
        if order_value == 'total':
            orders = Commande.objects.all().annotate(total_order=Sum(F('Cartdbs__qte') * F('Cartdbs__prix'))).order_by('-total_order')
        elif order_value == 'qte':
            orders = Commande.objects.all().annotate(qte_order=Count('Cartdbs')).order_by('-qte_order')
        else:
            orders = Commande.objects.all().order_by(order_value)
    else:
        orders = Commande.objects.all().order_by('-date', 'statut')

    if request.method == "GET":
        if form.is_valid():
            statut = form.cleaned_data['statut']
            clients = form.cleaned_data['clients']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            produits = form.cleaned_data['produits']
            especes = form.cleaned_data['especes']
            varietes = form.cleaned_data['varietes']
            portegreffes = form.cleaned_data['portegreffes']

            if statut.exists():
                orders = orders.filter(statut__in=statut)
            if clients.exists():
                orders = orders.filter(client__in=clients)
            if not start_date is None:
                orders = orders.filter(date__gte=start_date)
            if not end_date is None:
                orders = orders.filter(date__lte=end_date)
            if produits.exists():
                orders = orders.filter(Cartdbs__produit__in=produits)
            if especes.exists():
                orders = orders.filter(Cartdbs__produit__espece__in=especes)
            if varietes.exists():
                orders = orders.filter(Cartdbs__produit__variete__in=varietes)
            if portegreffes.exists():
                orders = orders.filter(Cartdbs__produit__portegreffe__in=portegreffes)
        else:
            orders = orders.filter(date__year__gte=datetime.now().year - 1)

        if 'max_val' in request.GET:
            try:
                max_value = int(request.GET['max_val'])
            except:
                max_value = 5
            orders = orders.filter(Cartdbs__qte__lte=max_value)

    context = {'formAction': formAction,
               'form': form,
               'orders': orders,
               'orders_list': orders,
               }
    return render(request, 'order/etiquettes.html', context)


@login_required
@staff_member_required
def print_etiquettes(request):
    if request.method == "POST":
        query_order_list = request.POST.getlist('checkorder')

    commandes = Commande.objects.filter(pk__in=query_order_list)
    orders = Cartdb.objects.filter(commande__in=commandes).order_by('produit__espece', 'produit__portegreffe', 'produit__variete')
    context = {
        'commandes': commandes,
        'orders': orders,
        'margin': id,
    }

    path_pdf = 'pdf/etiquettes.html'
    template = get_template(path_pdf)
    html = template.render(context)
    # pdf = render_to_pdf(path_pdf, context)
    pdf = False
    if pdf:
        # response = HttpResponse(pdf, content_type='application/pdf')
        response = HttpResponse(content_type='application/pdf')
        result = generate_pdf(path_pdf, file_object=response, context=context)

        # response = HttpResponse(html, content_type='text/html')
        today = date.today()
        filename = "Etiquettes_%s.pdf" % (today.strftime('%D%M%Y'))
        content = "inline; filename=%s" % filename
        download = request.POST["download"]
        download = False
        if download:
            content = "attachment; filename=%s" % filename
        response['Content-Disposition'] = content
        # return response
        return result
    return render(request, 'pdf/etiquettes.html', context)
    # return HttpResponse("Not found")


@login_required
@staff_member_required
def export_etiquettes(request):
    if request.method == "POST":
        query_order_list = request.POST.getlist('checkorder')
        # for order in request.POST.getlist('checkorder'):
        #     print(order)
        try:
            from io import BytesIO as IO
        except:
            from io import StringIO as IO

        excel_file = IO()
        xlwriter = pandas.ExcelWriter(excel_file, engine='xlsxwriter')

        workbook = xlwriter.book
        worksheet = workbook.add_worksheet("Etiquettes")

        format = workbook.add_format({
            'bold': 1,
            'align': 'center',
            'valign': 'vcenter',
            'font_size': 16
        })

        worksheet.set_column(0, 0, 25)
        worksheet.set_column(1, 1, 78)
        worksheet.set_default_row(57.75)
        i = 0
        row = 1
        commandes = Commande.objects.filter(pk__in=query_order_list)
        for commande in commandes:
            info_commande = str(commande.id) + " | " + commande.client.nom + " " + commande.client.prenom

            if commande.frais:
                info_commande += " | FRAIS : " + str(commande.frais)
            worksheet.write(row - 1, 1, info_commande, format)
            row += 1

            orders = Cartdb.objects.filter(commande=commande)
            for order in orders:
                worksheet.write(row - 1, 0, commande.id, format)
                info_order = str(order.qte) + " x " + order.produit.espece.nom + "-" + order.produit.variete.nom + "-" + order.produit.portegreffe.nom
                worksheet.write(row - 1, 1, info_order, format)
                row += 1

        xlwriter.save()
        xlwriter.close()
        excel_file.seek(0)
        filename = 'Etiquettes'
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="' + filename + '.xlsx"'

        return response


# *************************************************************************************
# DIVERS (TVA/FRAIS/STATUT)
# *************************************************************************************
@login_required
@staff_member_required
def manage_divers(request):
    frais = Frais.objects.all()
    tvas = Tva.objects.all()
    statuts = Statut.objects.all()

    previous_page = reverse('order:order-administration')
    context = {
        'frais': frais,
        'tvas': tvas,
        'statuts': statuts,
        'previous_page': previous_page,
    }
    return render(request, 'order/manage_divers.html', context)


# *************************************************************************************
# TVA (ADD/UPDATE/DELETE)
# *************************************************************************************
@login_required
@staff_member_required
def add_tva(request):
    if request.user.is_staff:
        title = "TVA"
        form = FormAddTva(request.POST or None)
        previous_page = reverse('order:manage-divers')
        formAction = 'order:add-tva'

        if request.method == 'POST':
            if form.is_valid():
                active = form.cleaned_data['active']
                if active is True:
                    tvas = Tva.objects.all()
                    for tva in tvas:
                        tva.active = False
                        tva.save()

                form.save()
                message = "Taux de TVA ajouté avec succès !"
                messages.success(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            return redirect('order:manage-divers')

        context_header = {
        }

        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "order/form_tva.html", context)
    else:
        return redirect('order:manage-divers')


@login_required
@staff_member_required
def edit_tva(request, tva_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "TVA"
        tva = Tva.objects.get(id=tva_id)
        form = FormAddTva(request.POST or None, instance=tva)

        previous_page = reverse('order:manage-divers')
        formAction = 'order:edit-tva', tva_id

        if request.POST:
            if form.is_valid():

                tva = float(form.cleaned_data['tva'])
                default = form.cleaned_data['default']

                if isinstance(tva, int) or isinstance(tva, float):

                    # SI EDITION (tva_id) et CHANGEMENT DE TVA ACTIVE
                    if default is True:
                        others_tva = Tva.objects.exclude(id=tva_id)
                        for other_tva in others_tva:
                            other_tva.default = False
                            other_tva.save()

                    message = "Taux de TVA modifié avec succès !"
                    messages.success(request, message)
                    instance = form.instance
                    obj = instance.save()

                else:
                    message = "Mauvaise saisie !"
                    messages.error(request, message)

                return redirect('order:manage-divers')
        context = {
            'tva_id': tva_id,
            'tva': tva,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "order/form_tva.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_tva(request, tva_id):
    if request.user.is_staff:
        tva = Tva.objects.get(id=tva_id)
        message = "Taux de TVA supprimé avec succès !"

        tva.delete()
        messages.success(request, message)
        return redirect('order:manage-divers')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def default_tva(request, tva_id):
    if request.user.is_staff:
        tva = Tva.objects.get(id=tva_id)
        message = "Taux de TVA passé en taux par défaut !"
        all_tva = Tva.objects.all()
        for item in all_tva:
            item.default = False
            item.save()
        tva.default = True
        tva.save()
        messages.success(request, message)
        return redirect('order:manage-divers')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# *************************************************************************************
# FRAIS
# *************************************************************************************
@login_required
@staff_member_required
def add_frais(request):
    if request.user.is_staff:
        title = "FRAIS"
        form = FormAddFrais(request.POST or None)
        previous_page = reverse('order:manage-divers')
        formAction = 'order:add-frais'

        if request.method == 'POST':
            if form.is_valid():
                obj = form.save()
                message = "Type de Frais ajouté avec succès !"
                messages.success(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            return redirect('order:manage-divers')

        context_header = {
        }

        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "order/form_frais.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_frais(request, frais_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "FRAIS"
        frais = Frais.objects.get(id=frais_id)
        form = FormAddFrais(request.POST or None, instance=frais)
        message = "Frais modifié avec succès !"

        previous_page = reverse('order:manage-divers')
        formAction = 'order:edit-frais', frais_id

        if request.POST:
            if form.is_valid():
                instance = form.instance
                obj = instance.save()
                messages.success(request, message)
                return redirect('order:manage-divers')
        context = {
            'frais_id': frais_id,
            'frais': frais,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "order/form_frais.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_frais(request, frais_id):
    if request.user.is_staff:
        title = "TVA"
        frais = Frais.objects.get(id=frais_id)
        form = FormAddTva(request.POST or None, instance=frais)
        message = "Frais supprimé avec succès !"

        frais.delete()
        messages.success(request, message)
        return redirect('order:manage-divers')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# *************************************************************************************
# STATUT
# *************************************************************************************
@login_required
@staff_member_required
def add_statut(request):
    if request.user.is_staff:
        title = "STATUT"
        form = FormAddStatut(request.POST or None)
        previous_page = reverse('order:manage-divers')
        formAction = 'order:add-statut'

        if request.method == 'POST':
            if form.is_valid():
                obj = form.save()
                message = "Type de Statut ajouté avec succès !"
                messages.success(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            return redirect('order:manage-divers')

        context_header = {
        }

        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "order/form_statut.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_statut(request, statut_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "STATUT"
        statut = Statut.objects.get(id=statut_id)
        form = FormAddStatut(request.POST or None, instance=statut)
        message = "Statut modifié avec succès !"

        previous_page = reverse('order:manage-divers')
        formAction = 'order:edit-statut', statut_id

        if request.POST:
            if form.is_valid():
                instance = form.instance
                obj = instance.save()
                messages.success(request, message)
                return redirect('order:manage-divers')
        context = {
            'frais_id': statut_id,
            'frais': statut,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "order/form_statut.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_statut(request, statut_id):
    if request.user.is_staff:
        title = "STATUT"
        statut = Statut.objects.get(id=statut_id)
        form = FormAddStatut(request.POST or None, instance=statut)
        message = "Statut supprimé avec succès !"

        statut.delete()
        messages.success(request, message)
        return redirect('order:manage-divers')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# ###########################################################################################################################################
# EXPORT
# ###########################################################################################################################################
@login_required
@staff_member_required
def export_commandes_xls(request):
    """
    Export Excel de l'ensemble des commandes du site
    """
    output = io.BytesIO()

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output)
    worksheet_commandes = workbook.add_worksheet("COMMANDES")
    worksheet_produits = workbook.add_worksheet("PRODUITS")

    cell_format_date = workbook.add_format()
    cell_format_date.set_num_format('dd/mm/yyyy hh:mm')

    # COMMANDES
    worksheet_commandes.write(0, 0, 'ID')
    worksheet_commandes.write(0, 1, 'DATE')
    worksheet_commandes.write(0, 2, 'CLIENT')
    worksheet_commandes.write(0, 3, 'REMISE')
    worksheet_commandes.write(0, 4, 'STATUT')
    worksheet_commandes.write(0, 5, 'TOTAL')
    worksheet_commandes.write(0, 6, 'DATE_UPDATE')
    worksheet_commandes.write(0, 7, 'TVA')
    worksheet_commandes.write(0, 8, 'FRAIS')
    worksheet_commandes.write(0, 9, 'FDP')
    worksheet_commandes.write(0, 10, 'INVENTAIRE')

    commandes = Commande.objects.all()
    row = 1
    for commande in commandes:
        worksheet_commandes.write(row, 0, commande.id)
        worksheet_commandes.write_datetime(row, 1, commande.date.replace(tzinfo=None), cell_format_date)
        worksheet_commandes.write(row, 2, commande.client.id)
        if not commande.remise is None:
            worksheet_commandes.write(row, 3, commande.remise)
        worksheet_commandes.write(row, 4, commande.statut.id)
        worksheet_commandes.write(row, 5, commande.total)
        if not commande.date_update is None:
            worksheet_commandes.write_datetime(row, 6, commande.date_update.replace(tzinfo=None), cell_format_date)
        if not commande.tva is None:
            worksheet_commandes.write(row, 7, commande.tva)
        if not commande.frais is None:
            worksheet_commandes.write(row, 8, commande.frais.id)
        if not commande.fdp is None:
            worksheet_commandes.write(row, 9, commande.fdp)
        row += 1

        # PRODUITS
        worksheet_produits.write(0, 0, 'ID')
        worksheet_produits.write(0, 1, 'QTE')
        worksheet_produits.write(0, 2, 'PRIX')
        worksheet_produits.write(0, 3, 'COMMANDE')
        worksheet_produits.write(0, 4, 'PRODUIT')

        produits = Cartdb.objects.all()
        row = 1
        for produit in produits:
            worksheet_produits.write(row, 0, produit.id)
            worksheet_produits.write(row, 1, produit.qte)
            worksheet_produits.write(row, 2, produit.prix)
            worksheet_produits.write(row, 3, produit.commande.id)
            worksheet_produits.write(row, 4, produit.produit.id)
            row += 1

    workbook.close()
    output.seek(0)

    filename = 'ExportCommandes.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


@login_required
@staff_member_required
def export_order_csv(request):
    if request.method == 'POST':
        categorie = request.POST.get('categorie')
        if categorie is None:
            message = "Catégorie manquante ... Veuillez réessayer !"
            messages.error(request, message)
            return redirect('order:export-order-xls')

        if categorie == "COMMANDES":
            produit_resource = CommandeResource()
            filename = "Commandes.csv"
        if categorie == "CLIENTS":
            produit_resource = ClientResource()
            filename = "Clients.csv"
        if categorie == "USERS":
            produit_resource = UserResource()
            filename = "Users.csv"
        if categorie == "TVA":
            produit_resource = TvaResource()
            filename = "Tva.csv"
        if categorie == "FRAIS":
            produit_resource = FraisResource()
            filename = "Frais.csv"
        if categorie == "STATUT":
            produit_resource = StatutResource()
            filename = "Statuts.csv"
        if categorie == "INVENTAIRES":
            produit_resource = InventaireResource()
            filename = "Periodes.csv"
        if categorie == "PRODUITS":
            produit_resource = ProduitsCommandeResource()
            filename = "ProduitsCommandes.csv"

        dataset = produit_resource.export()
        response = HttpResponse(dataset.csv, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename='+filename
        return response

    previous_page = reverse('order:order-administration')
    return render(request, 'order/export_order.html', {'previous_page': previous_page})


@login_required
@staff_member_required
def import_order_csv(request):
    previous_page = reverse('order:order-administration')
    if request.method == 'POST':
        categorie = request.POST.get('categorie')
        if categorie is None:
            message = "Catégorie manquante ... Veuillez réessayer !"
            messages.error(request, message)
            return redirect('order:import-order-xls')

        if categorie == "COMMANDES":
            produit_resource = CommandeResource()
        if categorie == "CLIENTS":
            produit_resource = ClientResource()
        if categorie == "USERS":
            produit_resource = UserResource()
        if categorie == "TVA":
            produit_resource = TvaResource()
        if categorie == "STATUT":
            produit_resource = StatutResource()
        if categorie == "FRAIS":
            produit_resource = FraisResource()
        if categorie == "INVENTAIRES":
            produit_resource = InventaireResource()
        if categorie == "PRODUITS":
            produit_resource = ProduitsCommandeResource()

        dataset = Dataset()
        new_datas = request.FILES['myfile']
        imported_data = dataset.load(new_datas.read().decode(), format='csv')
        result = produit_resource.import_data(dataset, dry_run=True)  # Test the data import

        if not result.has_errors():
            produit_resource.import_data(dataset, dry_run=False)  # Actually import now
            message = "Fichier importé avec succès !"
            messages.success(request, message)
            return redirect('order:order-administration')

    return render(request, 'order/import_order.html', {'previous_page': previous_page})


@login_required
@staff_member_required
def export_clients_xls(request):
    """
    Export Excel de l'ensemble des clients
    """
    output = io.BytesIO()

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output)
    worksheet_clients = workbook.add_worksheet("CLIENTS")
    worksheet_users = workbook.add_worksheet("USER")

    cell_format_date = workbook.add_format()
    cell_format_date.set_num_format('dd/mm/yyyy hh:mm')

    # CLIENTS
    worksheet_clients.write(0, 0, 'ID')
    worksheet_clients.write(0, 1, 'NOM')
    worksheet_clients.write(0, 2, 'PRENOM')
    worksheet_clients.write(0, 3, 'ADRESSE')
    worksheet_clients.write(0, 4, 'CP')
    worksheet_clients.write(0, 5, 'VILLE')
    worksheet_clients.write(0, 6, 'TEL')
    worksheet_clients.write(0, 7, 'MAIL')
    worksheet_clients.write(0, 8, 'COMMENTAIRE')
    worksheet_clients.write(0, 9, 'REMISE')
    worksheet_clients.write(0, 10, 'SOCIETE')
    worksheet_clients.write(0, 11, 'USER')

    clients = Client.objects.all()
    row = 1
    for client in clients:
        worksheet_clients.write(row, 0, client.id)
        worksheet_clients.write(row, 1, client.nom)
        worksheet_clients.write(row, 2, client.prenom)
        worksheet_clients.write(row, 3, client.adresse)
        worksheet_clients.write(row, 4, client.cp)
        worksheet_clients.write(row, 5, client.ville)
        worksheet_clients.write(row, 6, client.tel)
        worksheet_clients.write(row, 7, client.mail)
        worksheet_clients.write(row, 8, client.commentaire)
        if not client.remise is None:
            worksheet_clients.write(row, 9, client.remise)
        worksheet_clients.write(row, 10, client.societe)
        if not client.user is None:
            worksheet_clients.write(row, 11, client.user.id)
        row += 1

    # USERS
    worksheet_users.write(0, 0, 'ID')
    worksheet_users.write(0, 1, 'NOM')
    worksheet_users.write(0, 2, 'PRENOM')
    worksheet_users.write(0, 3, 'USERNAME')
    worksheet_users.write(0, 4, 'MAIL')
    worksheet_users.write(0, 5, 'SUPERUSER')
    worksheet_users.write(0, 6, 'STAFF')
    worksheet_users.write(0, 7, 'ACTIF')
    worksheet_users.write(0, 8, 'DATE CREATION')
    worksheet_users.write(0, 9, 'LAST LOGIN')

    row = 1
    users = User.objects.all()
    for user in users:
        worksheet_users.write(row, 0, user.id)
        if not user.last_name is None:
            worksheet_users.write(row, 1, user.last_name)
        if not user.first_name is None:
            worksheet_users.write(row, 2, user.first_name)
        worksheet_users.write(row, 3, user.username)
        worksheet_users.write(row, 4, user.email)
        worksheet_users.write(row, 5, user.is_superuser)
        worksheet_users.write(row, 6, user.is_staff)
        worksheet_users.write(row, 7, user.is_active)
        worksheet_users.write_datetime(row, 8, user.date_joined.replace(tzinfo=None), cell_format_date)
        if not user.last_login is None:
            worksheet_users.write_datetime(row, 9, user.last_login.replace(tzinfo=None), cell_format_date)
        row += 1

    workbook.close()
    output.seek(0)

    filename = 'ExportClients.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


@login_required
@staff_member_required
def export_divers_xls(request):
    """
    Export Excel de l'ensemble des informations annexes (TVA, STATUT, FRAIS)
    """
    output = io.BytesIO()

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output)
    worksheet_tva = workbook.add_worksheet("TVA")
    worksheet_frais = workbook.add_worksheet("FRAIS")
    worksheet_statut = workbook.add_worksheet("STATUT")
    worksheet_inventaire = workbook.add_worksheet("INVENTAIRE")

    cell_format_date = workbook.add_format()
    cell_format_date.set_num_format('dd/mm/yyyy hh:mm')

    # TVA
    worksheet_tva.write(0, 0, 'ID')
    worksheet_tva.write(0, 1, 'TAUX')
    worksheet_tva.write(0, 2, 'DEFAUT')
    worksheet_tva.write(0, 2, 'ACTIVE')

    tvas = Tva.objects.all()
    row = 1
    for tva in tvas:
        worksheet_tva.write(row, 0, tva.id)
        worksheet_tva.write(row, 1, tva.nom)
        worksheet_tva.write(row, 1, tva.default)
        worksheet_tva.write(row, 1, tva.active)
        row += 1

    # FRAIS
    worksheet_frais.write(0, 0, 'ID')
    worksheet_frais.write(0, 1, 'NOM')
    worksheet_frais.write(0, 2, 'ID TVA')
    worksheet_frais.write(0, 3, 'TVA')
    frais = Frais.objects.all()
    row = 1
    for f in frais:
        worksheet_frais.write(row, 0, f.id)
        worksheet_frais.write(row, 1, f.nom)
        worksheet_frais.write(row, 2, f.tva)
        worksheet_frais.write(row, 3, Tva.objects.get(id=f.tva))
        row += 1

        # STATUT
        worksheet_statut.write(0, 0, 'ID')
        worksheet_statut.write(0, 1, 'NOM')
        statuts = Statut.objects.all()
        row = 1
        for statut in statuts:
            worksheet_statut.write(row, 0, statut.id)
            worksheet_statut.write(row, 1, statut.nom)
            row += 1

        # INVENTAIRE
        worksheet_inventaire.write(0, 0, 'ID')
        worksheet_inventaire.write(0, 1, 'NOM')
        inventaires = Inventaire.objects.all()
        row = 1
        for inventaire in inventaires:
            worksheet_inventaire.write(row, 0, inventaire.id)
            worksheet_inventaire.write(row, 1, inventaire.start_date)
            worksheet_inventaire.write(row, 2, inventaire.end_date)
            row += 1

    workbook.close()
    output.seek(0)

    filename = 'ExportDivers.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


# ###########################################################################################################################################
# ADMINISTRATION
# ###########################################################################################################################################
@login_required
@staff_member_required
def order_administration(request):
    print(datetime.now())
    try:
        inventaire_actif = Inventaire.objects.get(actif=True)
    except:
        messages.error(request, "Veuillez créer une période de commande avant toute chose !")
        return redirect('order:manage-inventaire')

    pre_orders = Commande.objects.filter(statut__nom="Pré-commande")
    waiting_orders = Commande.objects.exclude(statut__nom__in=["Terminée", "Annulée", "Pré-commande"]).exclude(inventaire=inventaire_actif)
    context = {
        'pre_orders': pre_orders,
        'waiting_orders': waiting_orders,
    }
    if request.user.is_staff:
        return render(request, 'order/administration_menu_order.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def manage_order(request):

    if 'mo' in request.GET:
        order_value = request.GET['mo']
        if order_value == 'total':
            queryset = Commande.objects.all().annotate(total_order=Sum(F('Cartdbs__qte') * F('Cartdbs__prix'))).order_by('-total_order')
        else:
            queryset = Commande.objects.all().order_by(order_value)
        request.session['mo'] = request.GET['mo']
    elif 'mo' in request.session:
        order_value = request.session['mo']
        if order_value == 'total':
            queryset = Commande.objects.all().annotate(total_order=Sum(F('Cartdbs__qte') * F('Cartdbs__prix'))).order_by('-total_order')
        else:
            queryset = Commande.objects.all().order_by(order_value)
    else:
        queryset = Commande.objects.all().order_by('-date', 'statut')

    if request.user.is_staff:
        produits_commande = {}
        form = SearchOrderForm()

        if request.method == 'GET':
            form = SearchOrderForm(request.GET)
            if form.is_valid():
                statut = form.cleaned_data['statut']
                clients = form.cleaned_data['clients']
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']
                produits = form.cleaned_data['produits']
                especes = form.cleaned_data['especes']
                varietes = form.cleaned_data['varietes']
                portegreffes = form.cleaned_data['portegreffes']
                frais = form.cleaned_data['frais']
                inventaire = form.cleaned_data['inventaire']

                if produits.exists():
                    queryset = queryset.filter(Cartdbs__produit__in=produits)
                if especes.exists():
                    queryset = queryset.filter(Cartdbs__produit__espece__in=especes)
                if varietes.exists():
                    queryset = queryset.filter(Cartdbs__produit__variete__in=varietes)
                if portegreffes.exists():
                    queryset = queryset.filter(cCartdbs__produit__portegreffe__in=portegreffes)
                if statut.exists():
                    queryset = queryset.filter(statut__in=statut)
                if frais.exists():
                    queryset = queryset.filter(frais__in=frais)
                if clients.exists():
                    queryset = queryset.filter(client__in=clients)
                if not start_date is None:
                    queryset = queryset.filter(date__gte=start_date)
                if not end_date is None:
                    queryset = queryset.filter(date__lte=end_date)
                if inventaire.exists():
                    queryset = queryset.filter(inventaire__in=inventaire)
                else:
                    inventaire = Inventaire.objects.get(actif=True)
                    queryset = queryset.filter(inventaire=inventaire)
        if 'max_val' in request.GET:
            try:
                max_value = int(request.GET['max_val'])
            except:
                max_value = 5
            queryset = get_orders_items_max(max_value)

        for commande in queryset:
            produits = Cartdb.objects.filter(commande=commande)

            list_produit = []
            for produit in produits:
                dic_produit = {}
                dic_produit['id'] = produit.id
                dic_produit['qte'] = produit.qte
                dic_produit['produit'] = produit.produit
                dic_produit['prix'] = produit.prix
                list_produit.append(dic_produit)
            produits_commande[commande] = list_produit

        title = "Commandes"
        header = "Ajouter un Produit"
        javascript = "Cela va supprimer la commande"
        formAction = "order:manage-order"
        previous_page = reverse('order:order-administration')

        paginator = Paginator(queryset, 50)
        get_data = request.GET.copy()
        page = get_data.pop('page', None)
        mo = get_data.pop('mo', None)

        if 'page' in request.GET:
            page = request.GET['page']
        elif 'page' in request.session:
            page = request.session['page']

        try:
            commandes = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            commandes = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            commandes = paginator.page(paginator.num_pages)

        context_header = {
            'header': header,
            'javascript': javascript,
        }
        context = {
            'commandes': commandes,
            'produits_commande': produits_commande,
            'paginate': True,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'formAction': formAction,
            'form': form,
            'orders_list': queryset,
            'query_string': get_data.urlencode(),
        }

        return render(request, 'order/manage_order.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('onlineshop:produit-list')


@login_required
@staff_member_required
def edit_order(request, order_id):
    if request.GET.get('ano') == "1":
        request.session['ano'] = "1"
    else:
        request.session['ano'] = "0"

    if request.user.is_staff:
        title = "COMMANDES"
        order = Commande.objects.get(id=order_id)
        old_client = str(order.client.nom) + ' ' + str(order.client.prenom)
        old_remise = order.remise
        old_tva = order.tva.tva
        old_inventaire = order.inventaire.start_date.strftime('%Y') + '-' + order.inventaire.end_date.strftime('%Y')
        old_frais = ""
        old_montant_frais = ""
        old_statut = order.statut.nom
        if order.frais:
            old_frais = order.frais.nom
            old_montant_frais = order.montant_frais

        produits = Cartdb.objects.filter(commande=order)
        form = FormAddOrder(request.POST or None, instance=order)

        previous_page = reverse('order:manage-order')
        formAction = 'order:edit-order', order_id

        if request.POST:
            if form.is_valid():
                instance = form.instance
                obj = instance.save()

                new_client = str(instance.client.nom) + ' ' + str(instance.client.prenom)
                new_remise = instance.remise
                new_tva = instance.tva.tva
                new_inventaire = instance.inventaire.start_date.strftime('%Y') + '-' + instance.inventaire.end_date.strftime('%Y')
                new_frais = ""
                new_montant_frais = ""

                if instance.frais:
                    new_frais = instance.frais.nom
                    new_montant_frais = instance.montant_frais
                new_statut = instance.statut.nom

                if old_client != new_client:
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'client', old_client, new_client)
                if old_remise != new_remise:
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'remise', float(old_remise), float(new_remise))
                if old_tva != new_tva:
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'tva', float(old_tva), float(new_tva))
                if old_inventaire != new_inventaire:
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'inventaire', old_inventaire, new_inventaire)
                if old_frais != new_frais:
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'frais', old_frais, new_frais)
                if old_montant_frais != new_montant_frais:
                    if isinstance(old_montant_frais, float):
                        old_montant_frais = float(old_montant_frais)
                    else:
                        old_montant_frais = 0
                    if isinstance(new_montant_frais, float):
                        new_montant_frais = float(new_montant_frais)
                    else:
                        new_montant_frais = 0
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'montant_frais', old_montant_frais, new_montant_frais)
                if old_statut != new_statut:
                    log_order("Commande", str(request.user), order.pk, 'Edit', 'statut', old_statut, new_statut)

                order.date_update = datetime.now()
                order.save()

                message = "Commande modifiée avec succès !"
                messages.success(request, message)
                return redirect('order:manage-order')
            else:
                print(form.errors)
        context = {
            'order_id': order_id,
            'order': order,
            'produits': produits,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "order/form_order.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_order(request, order_id):
    if request.user.is_staff:
        title = "COMMANDES"
        order = Commande.objects.get(id=order_id)
        items = Cartdb.objects.filter(commande=order)

        for item in items:
            log_cart("Cart", str(request.user), item.pk, order.pk, item.produit.pk, 'Delete', 'qte', item.qte, 0)
            log_cart("Cart", str(request.user), item.pk, order.pk, item.produit.pk, 'Delete', 'prix', float(item.prix), '')
        #     produit = Produit.objects.get(id=item.produit.id)
        #     produit.stock_bis += item.qte
        #     produit.save()

        old_client = str(order.client.nom) + ' ' + str(order.client.prenom)
        old_remise = order.remise
        old_tva = order.tva.tva
        old_inventaire = order.inventaire.start_date.strftime('%Y') + '-' + order.inventaire.end_date.strftime('%Y')
        if order.frais:
            old_frais = order.frais.nom
            old_montant_frais = order.montant_frais
            log_order("Commande", str(request.user), order.pk, "Delete", 'frais', old_frais, '')
            log_order("Commande", str(request.user), order.pk, "Delete", 'montant_frais', float(old_montant_frais), '')
        old_statut = order.statut.nom

        log_order("Commande", str(request.user), order.pk, "Delete", 'client', old_client, '')
        log_order("Commande", str(request.user), order.pk, "Delete", 'remise', float(old_remise), '')
        log_order("Commande", str(request.user), order.pk, "Delete", 'tva', float(old_tva), '')
        log_order("Commande", str(request.user), order.pk, "Delete", 'inventaire', old_inventaire, '')
        log_order("Commande", str(request.user), order.pk, "Delete", 'statut', old_statut, '')

        order.delete()
        message = format_html("Commande supprimée avec succès !<br> <i class='bi bi-exclamation-triangle'></i> Stock NON remis à jour pour l'ensemble des produits de la commande")
        messages.success(request, message)
        return redirect('order:manage-order')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
def cancel_order(request, order_id):
    order = Commande.objects.get(id=order_id)
    items = Cartdb.objects.filter(commande=order)
    statut = Statut.objects.get(nom="Annulée")
    message = format_html("Commande annulée avec succès !")
    old_statut = order.statut.nom

    if request.user.is_staff:
        if order.statut.nom == "Terminée":
            message = message + format_html("<br>Stock final et virtuel remis à jour pour l'ensemble des produits de la commande")
            for item in items:
                old_final = item.produit.stock
                old_bis = item.produit.stock_bis

                item.produit.stock += item.qte
                item.produit.stock_bis += item.qte

                new_final = item.produit.stock
                new_bis = item.produit.stock_bis

                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Cancel", "sf", old_final, new_final)
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Cancel", "sb", old_bis, new_bis)

        if order.statut.nom == "Validée" or order.statut.nom == "En cours":
            message = message + format_html("<br>Stock virtuel remis à jour pour l'ensemble des produits de la commande")
            for item in items:
                old_bis = item.produit.stock_bis
                item.produit.stock_bis += item.qte
                new_bis = item.produit.stock_bis
                item.produit.save()
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Cancel", "sb", old_bis, new_bis)

        if order.statut.nom == "Pré-commande":
            message = message + format_html("<br>Stock futur remis à jour pour l'ensemble des produits de la commande")
            for item in items:
                old_future = item.produit.stock_future
                item.produit.stock_future -= item.qte
                item.produit.save()
                new_future = item.produit.stock_future
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Cancel", "sp", old_future, new_future)

        order.statut = statut
        order.date_update = datetime.now()
        order.save()

    new_statut = statut.nom
    log_order("Commande", str(request.user), order.pk, "Cancel", "statut", old_statut, new_statut)
    
    messages.success(request, message)
    return redirect('order:manage-order')


@login_required
@staff_member_required
def validate_order(request, order_id):
    if request.user.is_staff:
        title = "COMMANDES"
        order = Commande.objects.get(id=order_id)
        items = Cartdb.objects.filter(commande=order)
        statut = Statut.objects.get(nom="Validée")
        nb_produit = 0
        message = format_html("Commande validée avec succès !")

        old_statut = order.statut.nom
        
        # TERMINEE ___________________________________________________________________________________________________________
        if order.statut.nom == "Terminée":
            message = message + format_html("<br>Stock final remis à jour pour l'ensemble des produits de la commande.")
            for item in items:
                old_final = item.produit.stock
                item.produit.stock += item.qte
                new_final = item.produit.stock + item.qte
                item.produit.save()
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Valid", "sf", old_final, new_final)
                
        # ANNULEE OU EN ATTENTE ______________________________________________________________________________________________
        if order.statut.nom == "Annulée" or order.statut.nom == "En attente":
            message_produit = ""
            message = message + format_html("<br>Stock virtuel remis à jour pour l'ensemble des produits de la commande.")
            # On test avant si sur TOUS les produits le stock est OK !
            for item in items:
                if item.produit.stock_bis - item.qte < 0:
                    nb_produit += 1
                    message_produit = message_produit + "<li>" + item.produit.nom + "</li>"

            if nb_produit > 0:
                message = format_html("Impossible de passer la commande en \"Validée\" !<br>Stock insuffisant sur les produits suivants : <ul>" + message_produit + "</ul>")
                messages.error(request, message)
                return redirect('order:manage-order')

            for item in items:
                old_bis = item.produit.stock_bis
                new_bis = item.produit.stock_bis - item.qte

                item.produit.stock_bis -= item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Valid", "sb", old_bis, new_bis)

        # PRE COMMANDE ______________________________________________________________________________________________________
        if order.statut.nom == "Pré-commande":
            message_produit = ""
            message = message + format_html("<br>Stock futur et virtuel mis à jour pour l'ensemble des produits de la commande.")
            # On test avant si sur TOUS les produits le stock est OK !
            for item in items:
                if item.produit.stock_bis - item.qte < 0:
                    nb_produit += 1
                    message_produit = message_produit + "<li>" + item.produit.nom + "</li>"

            if nb_produit > 0:
                message = format_html("Impossible de passer la commande en \"Validée\" !<br>Stock insuffisant sur les produits suivants : <ul>" + message_produit + "</ul>")
                messages.error(request, message)
                return redirect('order:manage-order')

            for item in items:
                old_bis = item.produit.stock_bis
                old_future = item.produit.stock_future
                new_bis = item.produit.stock_bis - item.qte
                new_future = item.produit.stock_future - item.qte

                item.produit.stock_bis -= item.qte
                item.produit.stock_future -= item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Valid", "sb", old_bis, new_bis)
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Valid", "sp", old_future, new_future)

        order.statut = statut
        order.date_update = datetime.now()
        order.date_valid = datetime.now()
        order.save()

        new_statut = statut.nom
        log_order("Commande", str(request.user), order.pk, "Valid", "statut", old_statut, new_statut)
        
        messages.success(request, message)
        return redirect('order:manage-order')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def in_progress_order(request, order_id):
    if request.user.is_staff:
        title = "COMMANDES"
        order = Commande.objects.get(id=order_id)
        old_statut = order.statut.nom
        items = Cartdb.objects.filter(commande=order)
        statut = Statut.objects.get(nom="En cours")
        nb_produit = 0
        message = format_html("Commande passée en \"En cours\" avec succès !")

        if order.statut.nom == "Terminée":
            message = message + format_html("<br>Stock final remis à jour pour l'ensemble des produits de la commande.")
            for item in items:
                old_final = item.produit.stock
                new_final = item.produit.stock + item.qte

                item.produit.stock += item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Encours", "sf", old_final, new_final)

        if order.statut.nom == "Pré-commande":
            message_produit = ""
            message = message + format_html("<br>Stock future remis à jour pour l'ensemble des produits de la commande.")
            for item in items:
                if item.produit.stock_bis - item.qte < 0:
                    nb_produit += 1
                    message_produit = message_produit + "<li>" + item.produit.nom + "</li>"

            if nb_produit > 0:
                message = format_html("Impossible de passer la commande en \"En cours\" <br>Stock insuffisant sur les produits suivants : <ul>" + message_produit + "</ul>")
                messages.error(request, message)
                return redirect('order:manage-order')

            for item in items:
                # Desactivation de la mise à jours des stocks virtuels lors du passage d'une pré-commande vers une commande en cours
                old_bis = item.produit.stock_bis
                new_bis = item.produit.stock_bis - item.qte
                old_future = item.produit.stock_future
                new_future = item.produit.stock_future - item.qte

                item.produit.stock_bis -= item.qte
                item.produit.stock_future -= item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Encours", "sb", old_bis, new_bis)
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Encours", "sp", old_future, new_future)
            order.date = datetime.now()

        if order.statut.nom == "Annulée":
            message_produit = ""
            message = message + format_html("<br>Stock virtuel remis à jour pour l'ensemble des produits de la commande.")
            for item in items:
                if item.produit.stock_bis - item.qte < 0:
                    nb_produit += 1
                    message_produit = message_produit + "<li>" + item.produit.nom + "</li>"

            if nb_produit > 0:
                message = format_html("Impossible de passer la commande en \"En cours\" <br>Stock insuffisant sur les produits suivants : <ul>" + message_produit + "</ul>")
                messages.error(request, message)
                return redirect('order:manage-order')

            for item in items:
                old_bis = item.produit.stock_bis
                new_bis = item.produit.stock_bis - item.qte

                item.produit.stock_bis -= item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "Encours", "sb", old_bis, new_bis)

        order.date_update = datetime.now()
        order.statut = statut
        order.save()

        new_statut = statut.nom
        log_order("Commande", str(request.user), order.pk, "Encours", "statut", old_statut, new_statut)

        messages.success(request, message)
        return redirect('order:manage-order')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def finish_order(request, order_id):
    if request.user.is_staff:
        order = Commande.objects.get(id=order_id)
        old_statut = order.statut.nom
        items = Cartdb.objects.filter(commande=order)
        statut = Statut.objects.get(nom="Terminée")
        nb_produit = 0
        message = format_html("Commande terminée avec succès !")
        if order.statut.nom == "Validée" or order.statut.nom == "En cours":
            message_produit = ""
            for item in items:
                if item.produit.stock - item.qte < 0:
                    nb_produit += 1
                    message_produit = message_produit + "<li>" + item.produit.nom + "</li>"

            if nb_produit > 0:
                message = format_html("Impossible de repasser la commande en \"Terminée\"<br>Stock insuffisant sur les produits suivants : <ul>" + message_produit + "</ul>")
                messages.error(request, message)
                return redirect('order:manage-order')

            for item in items:
                old_final = item.produit.stock
                new_final = item.produit.stock - item.qte

                item.produit.stock -= item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "End", "sf", old_final, new_final)

        if order.statut.nom == "Annulée" or order.statut.nom == "En attente":
            message_produit = ""
            for item in items:
                if item.produit.stock_bis - item.qte < 0 or item.produit.stock - item.qte < 0:
                    nb_produit += 1
                    message_produit = message_produit + "<li>" + item.produit.nom + "</li>"

            if nb_produit > 0:
                message = format_html("Impossible de repasser la commande en \"Terminée\"<br>Stock insuffisant sur les produits suivants : <ul>" + message_produit + "</ul>")
                messages.error(request, message)
                return redirect('order:manage-order')

            for item in items:
                old_bis = item.produit.stock_bis
                old_final = item.produit.stock
                new_bis = item.produit.stock_bis - item.qte
                new_final = item.produit.stock - item.qte

                item.produit.stock -= item.qte
                item.produit.stock_bis -= item.qte
                item.produit.save()

                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "End", "sb", old_bis, new_bis)
                log_produit("Produit", str(request.user), item.produit.pk, order.pk, "End", "sf", old_final, new_final)

            message = message + format_html("<br> Stock final et virtuel mis à jour pour l'ensemble des produits de la commande")

        order.statut = statut
        order.date_update = datetime.now()
        order.save()

        new_statut = statut.nom
        log_order("Commande", str(request.user), order.pk, "End", "statut", old_statut, new_statut)

        messages.success(request, message)
        return redirect('order:manage-order')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def add_produit_order(request, order_id, manage):

    if request.user.is_staff:
        title = "PRODUIT"
        commande = Commande.objects.get(id=order_id)
        form = FormAddProduit(request.POST or None, initial={'commande': order_id}, order=commande)
        previous_page = reverse('order:edit-order', args=(order_id,))
        formAction = 'order:add-produit-order'


        if request.method == 'POST':
            if form.is_valid():
                produit = form.cleaned_data['produit']
                qte = form.cleaned_data['qte']
                prix = form.cleaned_data['prix']
                produit_commande = Cartdb.objects.filter(commande=commande, produit=produit)

                # IMPOSSIBLE DE MODIFIER UNE COMMANDE ANNULEE OU TERMINEE (AJOUT DE PRODUIT)
                if commande.statut.nom == "Annulée" or commande.statut.nom == "Terminée":
                    message = format_html("Impossible de modifier les produits sur cette commande (problème de statut) !")
                    messages.error(request, message)
                    if manage == "1":
                        return redirect('order:edit-order', order_id)
                    else:
                        return redirect('order:order-detail', order_id)

                # COMMANDE EN COURS OU VALIDEE
                if not commande.statut.nom == "Pré-commande":
                    if produit_commande.exists():
                        produit = Produit.objects.get(pk=produit.id)
                        produit_commande = produit_commande.first()
                        if produit.stock_bis - qte >= 0:
                            old_cart_qte = produit_commande.qte
                            old_cart_prix = produit_commande.prix

                            produit_commande.qte = produit_commande.qte + qte
                            produit_commande.prix = prix
                            produit_commande.save()

                            new_cart_qte = produit_commande.qte + qte
                            new_cart_prix = prix

                            if old_cart_prix != new_cart_prix:
                                log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Edit", "prix", float(old_cart_prix), float(new_cart_prix))
                            log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Edit", "qte", old_cart_qte, new_cart_qte)

                            old_bis = produit.stock_bis
                            new_bis = produit.stock_bis - qte

                            produit.stock_bis = produit.stock_bis - qte
                            produit.save()

                            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sb", old_bis, new_bis)

                            message = format_html("Produit déjà présent dans la commande.<br/>Quantité et Prix mis à jour !")
                            messages.success(request, message)
                        else:
                            message = format_html("Produit déjà présent dans la commande.<br/>Stock insuffisant !")
                            messages.error(request, message)

                        if manage == "1":
                            return redirect('order:edit-order', order_id)
                        else:
                            return redirect('order:order-detail', order_id)
                    else:
                        if produit.stock_bis >= qte:
                            obj = form.save(commit=False)

                            new_qte = form.cleaned_data['qte']
                            new_prix = form.cleaned_data['prix']

                            obj.commande = commande
                            obj.save()

                            log_cart("Cart", str(request.user), obj.pk, commande.pk, produit.pk, "Add", "qte", 0, new_qte)
                            log_cart("Cart", str(request.user), obj.pk, commande.pk, produit.pk, "Add", "prix", '', float(new_prix))

                            old_bis = produit.stock_bis
                            new_bis = produit.stock_bis - qte

                            produit.stock_bis -= qte
                            produit.save()

                            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Add", "sb", old_bis, new_bis)

                            message = "Produit ajouté à la commande avec succès !"
                            messages.success(request, message)
                        else:
                            message = "Stock insuffisant !"
                            messages.error(request, message)
                # PRE-COMMANDE
                else:
                    if produit_commande.exists():
                        produit = Produit.objects.get(pk=produit.id)
                        produit_commande = produit_commande.first()

                        old_cart_qte = produit_commande.qte
                        old_cart_prix = produit_commande.prix
                        new_cart_qte = produit_commande.qte + qte
                        new_cart_prix = prix
                        produit_commande.qte = produit_commande.qte + qte
                        produit_commande.prix = prix
                        produit_commande.save()

                        if old_cart_prix != new_cart_prix:
                            log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Edit", "prix", float(old_cart_prix), float(new_cart_prix))
                        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Edit", "qte", old_cart_qte, new_cart_qte)

                        old_future = produit.stock_future
                        new_future = produit.stock_future + qte

                        produit.stock_future += qte
                        produit.save()

                        log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sp", old_future, new_future)

                        message = format_html("Produit déjà présent dans la pré-commande.<br/>Quantité et Prix mis à jour !")
                        messages.success(request, message)

                        if manage == "1":
                            return redirect('order:edit-order', order_id)
                        else:
                            return redirect('order:order-detail', order_id)
                    else:
                        obj = form.save(commit=False)
                        obj.commande = commande
                        obj.save()

                        new_cart_qte = qte
                        new_cart_prix = prix
                        log_cart("Cart", str(request.user), obj.pk, commande.pk, produit.pk, "Add", "prix", '', float(new_cart_prix))
                        log_cart("Cart", str(request.user), obj.pk, commande.pk, produit.pk, "Add", "qte", 0, new_cart_qte)

                        old_future = produit.stock_future
                        new_future = produit.stock_future + qte
                        produit.stock_future += qte
                        produit.save()

                        log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sp", old_future, new_future)

                    message = "Produit ajouté à la pré-commande avec succès !"
                    messages.success(request, message)

                commande.date_update = datetime.now()
                commande.save()

            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)

            if manage == "1":
                return redirect('order:edit-order', order_id)
            else:
                return redirect('order:order-detail', order_id)
        context_header = {
        }

        context = {
            'commande': commande,
            'form': form,
            'order_id': order_id,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'manage': manage,
        }
        return render(request, "order/form_produit.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_produit_order(request, order_id, produit_id):
    try:
        anomalie = request.session['ano']
    except:
        request.session['ano'] = "0"
        anomalie = "0"

    admin_mode = get_admin_mode(request.user)
    if request.user.is_staff:
        title = "PRODUIT"
        produit = Produit.objects.get(id=produit_id)
        commande = Commande.objects.get(id=order_id)
        produit_commande = Cartdb.objects.get(commande=commande, produit=produit)
        form = FormEditProduit(produit_id, request.POST or None, instance=produit_commande)
        form.fields['produit'].initial = produit
        previous_page = reverse('order:edit-order', args=(order_id,))
        formAction = 'order:edit-produit-order'
        previous_qte = produit_commande.qte

        old_bis = produit.stock_bis
        old_final = produit.stock
        old_future = produit.stock_future
        old_qte = produit_commande.qte
        old_prix = produit_commande.prix

        if request.method == 'POST':
            if form.is_valid():
                qte = form.cleaned_data['qte']
                prix = form.cleaned_data['prix']
                produit = form.cleaned_data['produit']

                if (produit.stock_bis + previous_qte) >= qte or admin_mode or anomalie == "1":
                    obj = form.save(commit=False)
                    obj.commande = commande
                    obj.save()

                    log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Update", "qte", old_qte, qte)
                    if prix != old_prix:
                        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Update", "prix", float(old_prix), float(prix))

                    # SI COMMANDE TERMINEE ON MET A JOUR LE STOCK FINAL EN PLUS DU STOCK VIRTUEL
                    if commande.statut.nom == "Terminée":
                        old_final = produit.stock
                        qte_to_modify_final = produit.stock + previous_qte - qte
                        produit.stock = qte_to_modify_final
                        log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sf", old_final, qte_to_modify_final)

                        commande.date_update = datetime.now()
                        commande.save()

                    # SI COMMANDE AUTRE QUE ANNULEE ON MET A JOUR LE STOCK VIRTUEL SINON RIEN
                    if commande.statut.nom != "Annulée" and commande.statut.nom !="Pré-commande" and anomalie != "1":
                        qte_to_modify = produit.stock_bis + previous_qte - qte
                        if qte_to_modify <= produit.stock:
                            old_bis = produit.stock_bis
                            produit.stock_bis = qte_to_modify
                            produit.save()
                            new_stock = qte_to_modify
                            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sb", old_bis, qte_to_modify)

                            commande.date_update = datetime.now()
                            commande.save()

                    if commande.statut.nom == "Pré-commande":
                        old_future = produit.stock_future
                        qte_to_modify = produit.stock_future - previous_qte + qte
                        produit.stock_future = qte_to_modify
                        produit.save()
                        log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sp", old_future, qte_to_modify)

                        commande.date_update = datetime.now()
                        commande.save()

                    message = "Produit de la commande édité avec succès !"
                    if anomalie == "1":
                        message = message + " (Stock non mis à jour car commande en anomalie)"
                    messages.success(request, message)

                elif previous_qte > qte and qte <= total_qte_inventaire_progress(produit) or admin_mode:
                    obj = form.save(commit=False)
                    obj.commande = commande
                    obj.save()

                    # SI COMMANDE TERMINEE ON MET A JOUR LE STOCK FINAL EN PLUS DU STOCK VIRTUEL
                    if commande.statut.nom == "Terminée":
                        old_final = produit.stock
                        qte_to_modify_final = produit.stock + previous_qte - qte
                        produit.stock = qte_to_modify_final
                        produit.save()
                        log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sf", old_final, qte_to_modify_final)

                        commande.date_update = datetime.now()
                        commande.save()

                    # SI COMMANDE AUTRE QUE ANNULEE ON MET A JOUR LE STOCK VIRTUEL SINON RIEN
                    if commande.statut.nom not in ["Annulée", "Terminée", "En attente"]:
                        old_bis = produit.stock_bis
                        qte_to_modify = produit.stock_bis + previous_qte - qte
                        if qte_to_modify <= produit.stock:
                            produit.stock_bis = qte_to_modify
                            produit.save()
                            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sb", old_bis, qte_to_modify)

                            commande.date_update = datetime.now()
                            commande.save()

                    log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.ok, "Update", "qte", old_qte, qte)
                    if prix != old_prix:
                        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.ok, "Update", "prix", float(old_prix), float(prix))

                elif commande.statut.nom == "Pré-commande":
                    obj = form.save(commit=False)
                    obj.commande = commande
                    obj.save()
                    old_future = produit.stock_future
                    qte_to_modify = produit.stock_future - previous_qte + qte
                    produit.stock_future = qte_to_modify
                    produit.save()
                    log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sp", old_future, qte_to_modify)

                    commande.date_update = datetime.now()
                    commande.save()

                    message = "Produit de la pré-commande édité avec succès !"
                    messages.success(request, message)

                    log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.ok, "Update", "qte", old_qte, qte)
                    if prix != old_prix:
                        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.ok, "Update", "prix", float(old_prix), float(prix))
                else:
                    message = "Stock insuffisant !"
                    messages.error(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            # return redirect('order:edit-order', order_id)
            return custom_redirect('order:edit-order', order_id, ano=anomalie)

        context_header = {
        }

        context = {
            'anomalie': anomalie,
            'form': form,
            'order_id': order_id,
            'produit_id': produit_id,
            'produit': produit,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "order/form_produit.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_produit_order(request, order_id, produit_id):
    if request.user.is_staff:
        commande = Commande.objects.get(id=order_id)
        produit = Produit.objects.get(id=produit_id)
        old_final = produit.stock
        old_bis = produit.stock_bis
        old_future = produit.stock_future

        produit_commande = Cartdb.objects.get(commande=commande, produit=produit)
        # SI COMMANDE TERMINEE ON MET A JOUR LE STOCK FINAL EN PLUS DU STOCK VIRTUEL
        if commande.statut.nom == "Terminée":
            produit.stock += produit_commande.qte
            new_final = produit.stock + produit_commande.qte
            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sf", old_final, new_final)

        if commande.statut.nom == "Pré-commande":
            produit.stock_future -= produit_commande.qte
            new_future = produit.stock_future - produit_commande.qte
            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sp", old_future, new_future)

        # SI COMMANDE AUTRE QUE ANNULEE ON MET A JOUR LE STOCK VIRTUEL SINON RIEN
        if commande.statut.nom == "En cours" or commande.statut.nom == "Validée":
            produit.stock_bis += produit_commande.qte
            new_bis = produit.stock_bis
            log_produit("Produit", str(request.user), produit.pk, commande.pk, "Update", "sb", old_bis, new_bis)
        produit.save()
        old_cart_qte = produit_commande.qte
        old_cart_prix = produit_commande.prix

        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Delete", "qte", old_cart_qte, 0)
        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Delete", "prix", float(old_cart_prix), '')

        produit_commande.delete()



        commande.date_update = datetime.now()
        commande.save()
        message = "Produit supprimé de la commande avec succès et stock remis à jour !"
        messages.success(request, message)
        return redirect('order:edit-order', order_id)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def recycle_produit_order(request, order_id, produit_id):
    if request.user.is_staff:

        commande = Commande.objects.get(id=order_id)
        produit = Produit.objects.get(id=produit_id)
        produit_commande = Cartdb.objects.get(commande=commande, produit=produit)
        old_cart_qte = produit_commande.qte
        old_cart_prix = produit_commande.prix

        # ATTENTION : ON NE MET PAS A JOUR LES STOCKS
        produit_commande.delete()
        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Delete", "qte", old_cart_qte, 0)
        log_cart("Cart", str(request.user), produit_commande.pk, commande.pk, produit.pk, "Delete", "prix", float(old_cart_prix), '')

        commande.date_update = datetime.now()
        commande.save()

        message = format_html("Produit supprimé de la commande avec succès !<br><i class='bi bi-exclamation-triangle'></i> Stock non mis à jour ! <i class='bi bi-exclamation-triangle'></i>")
        messages.success(request, message)
        return redirect('order:edit-order', order_id)

    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def get_produit_stock(request, produit_id):
    produit = get_object_or_404(Produit, id=produit_id)
    print("Check :", produit.stock_bis)
    return JsonResponse({"total": produit.stock_bis})


# IMPORT DES PRODUITS A PARTIR D'UN FICHIER EXCEL
@login_required
@staff_member_required
def import_commandes_xls(request):
    previous_page = reverse('order:order-administration')

    try:
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            message = ""

            wb = load_workbook(filename=settings.CONTENT_DIR + excel_file, read_only=True)
            ws = wb['COMMANDES']
            max_col = ws.max_column
            max_row = ws.max_row

            # REMOVE DATA FROM TABLE COMMANDES
            if request.POST.get('delete_data', True):
                pass
                Commande.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Commande.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    date=ws.cell(row=i, column=2).value,
                    client=get_object_from_id(ws.cell(row=i, column=3).value, 'client'),
                    remise=ws.cell(row=i, column=4).value,
                    statut=get_object_from_id(ws.cell(row=i, column=5).value, 'statut'),
                    date_update=ws.cell(row=i, column=6).value,
                    tva=get_object_from_id(ws.cell(row=i, column=7).value, 'tva'),
                    frais=get_object_from_id(ws.cell(row=i, column=8).value, 'frais'),
                    montant_frais=ws.cell(row=i, column=9).value,
                    inventaire=get_object_from_id(ws.cell(row=i, column=8).value, 'inventaire'),
                )
                obj.save()
            print(max_row, " Commandes importées")
            message = "Fichier de commandes importé avec succès (" + str(max_row) + " commandes importés)<br/>"
            ws = wb['PRODUITS']
            max_col = ws.max_column
            max_row = ws.max_row

            # REMOVE DATA FROM TABLE CARTDB
            if request.POST.get('delete_data', True):
                Cartdb.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Cartdb.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    qte=ws.cell(row=i, column=2).value,
                    prix=ws.cell(row=i, column=3).value,
                    commande=get_object_from_id(ws.cell(row=i, column=4).value, 'commande'),
                    produit=get_object_from_id(ws.cell(row=i, column=5).value, 'produit'),
                )
                obj.save()
            print(max_row, " Produits associés importés")
            message = message + "Fichier de produits liés importé avec succès (" + str(max_row) + " produits liés importés)"
            messages.success(request, message)
            wb.close()

            return render(request, 'order/import_commande.html', {'previous_page': previous_page})

    except Exception as identifier:
        message = format_html("Une erreur est survenue ... merci de réessayer !<br>[ <i>" + str(identifier) + "</i> ]")
        messages.error(request, message)

    return render(request, 'order/import_commande.html', {'previous_page': previous_page})
    # return response


@login_required
@staff_member_required
def import_clients_xls(request):
    previous_page = reverse('order:order-administration')

    try:
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            message = ""
            wb = load_workbook(filename=settings.CONTENT_DIR + excel_file, read_only=True)

            # CLIENTS ----------------------------------------------------------------------------
            ws = wb['USER']
            max_col = ws.max_column
            max_row = ws.max_row

            # REMOVE DATA FROM TABLE USER
            if request.POST.get('delete_data', True):
                User.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = User.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    last_name=ws.cell(row=i, column=2).value,
                    first_name=ws.cell(row=i, column=3).value,
                    username=ws.cell(row=i, column=4).value,
                    email=ws.cell(row=i, column=5).value,
                    is_superuser=ws.cell(row=i, column=6).value,
                    is_staff=ws.cell(row=i, column=7).value,
                    is_active=ws.cell(row=i, column=8).value,
                    date_joined=ws.cell(row=i, column=9).value,
                    last_login=ws.cell(row=i, column=10).value,
                )
                obj.save()

            message = format_html("Fichier de Clients importé avec succès (" + str(max_row) + " users liés importés)<br>")
            # CLIENTS ----------------------------------------------------------------------------
            ws = wb['CLIENTS']
            max_col = ws.max_column
            max_row = ws.max_row

            # REMOVE DATA FROM TABLE CLIENTS
            if request.POST.get('delete_data', True):
                pass
                Client.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Client.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    nom=ws.cell(row=i, column=2).value,
                    prenom=ws.cell(row=i, column=3).value,
                    adresse=ws.cell(row=i, column=4).value,
                    cp=ws.cell(row=i, column=5).value,
                    ville=ws.cell(row=i, column=6).value,
                    tel=ws.cell(row=i, column=7).value,
                    mail=ws.cell(row=i, column=8).value,
                    commentaire=ws.cell(row=i, column=9).value,
                    remise=ws.cell(row=i, column=10).value,
                    user=get_object_from_id(ws.cell(row=i, column=11).value, 'user'),
                )
                obj.save()

            message = message + format_html("Fichier de clients importé avec succès (" + str(max_row) + " clients importés)<br/>")
            messages.success(request, message)
            wb.close()

            return render(request, 'order/import_clients.html', {'previous_page': previous_page})

    except Exception as identifier:
        message = format_html("Une erreur est survenue ... merci de réessayer !<br>[ <i>" + str(identifier) + "</i> ]")
        messages.error(request, message)

    return render(request, 'order/import_clients.html', {'previous_page': previous_page})


@login_required
@staff_member_required
def import_divers_xls(request):
    previous_page = reverse('order:order-administration')

    try:
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            message = format_html("Fichier Divers importé avec succès<br>")
            wb = load_workbook(filename=settings.CONTENT_DIR + excel_file, read_only=True)

            ws = wb['TVA']
            max_col = ws.max_column
            max_row = ws.max_row
            # REMOVE DATA FROM TABLE TVA
            if request.POST.get('delete_data', True):
                Tva.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Tva.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    tva=ws.cell(row=i, column=2).value,
                    default=ws.cell(row=i, column=3).value,
                    active=ws.cell(row=i, column=4).value,
                )
                obj.save()
            message += format_html("- " + str(max_row) + " Taux de TVA importés : <br/>")

            ws = wb['FRAIS']
            max_col = ws.max_column
            max_row = ws.max_row
            # REMOVE DATA FROM TABLE FRAIS
            if request.POST.get('delete_data', True):
                Frais.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Frais.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    nom=ws.cell(row=i, column=2).value,
                    tva=Tva.objects.get(id=ws.cell(row=i, column=3).value),
                )
                obj.save()
            message += message + format_html("- " + str(max_row) + " Frais importés<br/>")

            ws = wb['STATUT']
            max_col = ws.max_column
            max_row = ws.max_row

            # REMOVE DATA FROM TABLE STATUT
            if request.POST.get('delete_data', True):
                Statut.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Statut.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    nom=ws.cell(row=i, column=2).value,
                )
                obj.save()
            message = message + format_html("F- " + str(max_row) + " Statuts importés<br>")

            ws = wb['INVENTAIRE']
            max_col = ws.max_column
            max_row = ws.max_row

            # REMOVE DATA FROM TABLE INVENTAIRE
            if request.POST.get('delete_data', True):
                Inventaire.objects.all().delete()

            for i in range(2, max_row + 1):
                obj = Inventaire.objects.create(
                    id=ws.cell(row=i, column=1).value,
                    start_date=ws.cell(row=i, column=2).value,
                    end_date=ws.cell(row=i, column=3).value,
                )
                obj.save()
            message = message + format_html("F- " + str(max_row) + " Période importées<br>")

            messages.success(request, message)
            wb.close()

            return render(request, 'order/import_divers.html', {'previous_page': previous_page})

    except Exception as identifier:
        message = format_html("Une erreur est survenue ... merci de réessayer !<br>[ <i>" + str(identifier) + "</i> ]")
        messages.error(request, message)

    return render(request, 'order/import_divers.html', {'previous_page': previous_page})


# ----------------------------------------------------------------------------------------------------
# CLIENTS
# ----------------------------------------------------------------------------------------------------
@login_required
@staff_member_required
def manage_client(request):

    if request.user.is_staff:
        form = SearchClientForm()
        order_value = 'nom'
        if 'mc' in request.GET:
            order_value = request.GET['mc']
            if order_value == 'total':
                queryset = Client.objects.all().annotate(total_order=Count('Commandes')).order_by('-total_order')
            else:
                queryset = Client.objects.all().order_by(order_value)
            request.session['mc'] = request.GET['mc']
        elif 'mc' in request.session:
            order_value = request.session['mc']
            if order_value == 'total':
                queryset = Client.objects.all().annotate(total_order=Count('Commandes')).order_by('-total_order')
            else:
                queryset = Client.objects.all().order_by(order_value)
        else:
            queryset = Client.objects.all().order_by('nom', 'prenom')

        if request.method == 'GET':
            form = SearchClientForm(request.GET)
            if form.is_valid():
                cp = form.cleaned_data['cp']
                mail = form.cleaned_data['mail']
                ville = form.cleaned_data['ville']
                remise = form.cleaned_data['remise']
                nom = form.cleaned_data['nom']
                prenom = form.cleaned_data['prenom']
                activate = form.cleaned_data['activate']

                queryset = queryset.filter(nom__contains=nom)
                queryset = queryset.filter(mail__contains=mail)
                queryset = queryset.filter(prenom__contains=prenom)
                if activate == "2":
                    queryset = queryset.filter(Q(activate=True) | Q(activate=None))
                if activate == "3":
                    queryset = queryset.filter(activate=False)
                if not cp is None and cp != "":
                    queryset = queryset.filter(cp__startswith=cp)
                if not ville is None and len(ville) > 0:
                    queryset = queryset.filter(ville=ville)
                if not remise is None:
                    queryset = queryset.filter(remise=remise)

        count_order = {}
        dic_order = {}
        for client in queryset:
            dic_order[client.id] = list(Commande.objects.filter(client=client))
            count_order[client.id] = Commande.objects.filter(client=client).count()
        title = "Clients"
        header = "Ajouter un Client"
        javascript = "Cela va désactiver (Suppression impossible)"
        formAction = "order:manage-client"
        previous_page = reverse('order:order-administration')

        paginator = Paginator(queryset, 50)
        get_data = request.GET.copy()
        page = get_data.pop('page', None)
        mc = get_data.pop('mc', None)

        if 'page' in request.GET:
            page = request.GET['page']
        elif 'page' in request.session:
            page = request.session['page']

        try:
            clients = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            clients = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            clients = paginator.page(paginator.num_pages)

        context_header = {
            'header': header,
            'javascript': javascript,
        }
        context = {
            'dic_order': dic_order,
            'count_order': count_order,
            'clients': clients,
            'paginate': True,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'formAction': formAction,
            'form': form,
            'nb_clients': len(queryset),
            'query_string': get_data.urlencode(),
        }

        return render(request, 'order/manage_client.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('onlineshop:produit-list')


@login_required
@staff_member_required
def add_client(request):
    if request.user.is_staff:
        title = "CLIENTS"
        form = FormAddClient(request.POST or None)
        previous_page = reverse('order:manage-client')
        formAction = 'order:add-client'

        if request.method == 'POST':
            if form.is_valid():
                obj = form.save()

                user_mail = '{}.{}@mail.com'.format(obj.prenom.lower(), obj.nom.lower())
                if obj.mail == "":
                    obj.mail = user_mail
                    obj.save()

                user_mail = obj.mail
                username = '{}.{}'.format(obj.prenom.lower(), obj.nom.lower())

                test_username = User.objects.filter(username__startswith=username)
                if len(test_username) > 0:
                    username = '{}.{}_{}'.format(obj.prenom.lower(), obj.nom.lower(), len(test_username))

                user = User.objects.create_user(username, user_mail, '123456')
                user.last_name = obj.nom
                user.first_name = obj.prenom
                user.is_staff = False
                user.is_superuser = False
                obj_user = user.save()

                obj.user = user
                obj.save()

                message = "Client et User créé avec succès !"
                messages.success(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            return redirect('order:manage-client')

        context_header = {
        }

        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "order/form_client.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_client(request, client_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "CLIENT"
        client = Client.objects.get(id=client_id)
        print(client)
        try:
            user = User.objects.get(id=client.user.id)
            print(user)
        except:
            user = None

        form = FormAddClient(request.POST or None, instance=client, initial={'user': user})

        previous_page = reverse('order:manage-client')
        formAction = 'order:edit-client', client_id
        message = ""
        if request.POST:
            if form.is_valid():
                instance = form.instance
                print(instance.mail)
                email_exist = User.objects.filter(email=instance.mail)
                if len(email_exist) == 0:
                    obj = instance.save()
                    message = "Client modifié avec succès !"

                if not user is None:
                    user.first_name = instance.prenom
                    user.last_name = instance.nom
                    user.email = instance.mail
                    user.save()
                    message = "Client (et User) modifié avec succès !"

                messages.success(request, message)
                return redirect('order:manage-client')
        context = {
            'client_id': client_id,
            'client': client,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "order/form_client.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def activate_client(request, client_id):
    if request.user.is_staff:
        client = Client.objects.get(id=client_id)
        client.activate = True
        client.save()

        messages.success(request, "Client réactivé avec succès !")
        return redirect('order:manage-client')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_client(request, client_id):
    if request.user.is_staff:
        client = Client.objects.get(id=client_id)

        count = Commande.objects.filter(client=client).count()
        if count > 0:
            message = "Client desactivé avec succès ! (Le client a passer des commandes)"
            client.activate = False
            client.save()
        else:
            message = "Client supprimé avec succès ! (Le client n'avait pas encore passé de commande)"
            if not client.user is None:
                user = User.objects.get(id=client.user.id)
            client.delete()
            if not client.user is None:
                user.delete()
        messages.success(request, message)
        return redirect('order:manage-client')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def manage_inventaire(request):
    if request.user.is_staff:
        inventaires = Inventaire.objects.all().order_by('-start_date')
        context = {
            'inventaires': inventaires
        }
        return render(request, "order/manage_inventaire.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def add_inventaire(request):
    if request.user.is_staff:
        form = FormInventaire(request.POST or None)
        previous_page = reverse('order:manage-inventaire')
        formAction = 'order:add-inventaire'

        if request.method == 'POST':
            if form.is_valid():
                obj = form.save()
                message = "Période ajoutée avec succès !"
                messages.success(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            return redirect('order:manage-inventaire')

        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page
        }
        return render(request, "order/form_inventaire.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def actif_inventaire(request, inventaire_id):
    if request.user.is_staff:
        try:
            inventaire = Inventaire.objects.get(pk=inventaire_id)
        except:
            messages.error(request, "La période selectionnée n'existe pas !")
            return redirect('order:manage-inventaire')

        inventaire.actif = True
        inventaire.save()

        inventaires = Inventaire.objects.exclude(pk=inventaire_id)
        for inv in inventaires:
            inv.actif = False
            inv.save()

        messages.success(request, "Période activée avec succès")
        return redirect('order:manage-inventaire')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_inventaire(request, inventaire_id):
    if request.user.is_staff:
        try:
            inventaire = Inventaire.objects.get(pk=inventaire_id)
        except:
            messages.error(request, "La période selectionnée n'existe pas !")
            return redirect('order:manage-inventaire')

        form = FormInventaire(request.POST or None, instance=inventaire)
        previous_page = reverse('order:manage-inventaire')
        formAction = 'order:edit-inventaire'

        if request.POST:
            if form.is_valid():
                instance = form.instance
                obj = instance.save()
                messages.success(request, "Période modifiée avec succès")
                return redirect('order:manage-inventaire')

        context = {
            'inventaire': inventaire,
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
        }
        return render(request, "order/form_inventaire.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_inventaire(request, inventaire_id):
    if request.user.is_staff:
        return redirect('order:manage-inventaire')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def reset_order(request):
    cancel_statut = Statut.objects.get(nom='Annulée')
    done_statut = Statut.objects.get(nom='Terminée')
    inprogress_statut = Statut.objects.get(nom='En cours')
    valid_statut = Statut.objects.get(nom='Validée')
    wait_statut = Statut.objects.get(nom='En attente')

    form = FormResetOrder(request.POST or None)
    formAction = 'order:reset-order'
    previous_page = 'order:order-administration'
    inventaires = Inventaire.objects.all()
    commandes = Commande.objects.filter(statut__in=[inprogress_statut, valid_statut])

    if request.POST:
        if form.is_valid():
            inventaire = form.cleaned_data['inventaire']
            mode = form.cleaned_data['mode']
            commandes = commandes.filter(inventaire=inventaire, statut__in=[inprogress_statut, valid_statut])
            nb_produit = 0
            produit_list = ""
            if mode == "FULL":
                for commande in commandes:
                    # Modification du statut de la commande + changement de date (MAJ)
                    old_statut = commande.statut.nom
                    commande.statut = done_statut
                    commande.date_update = datetime.now()
                    commande.save()

                    log_order("Commande", str(request.user), commande.pk, "End", "statut", old_statut, commande.statut.nom)

                message = format_html("Les commandes en cours et validées ont bien toutes étés <b>Terminées</b> pour la période selectionnée !")
                messages.success(request, message)

            elif mode == "CHECK":
                for commande in commandes:
                    items = Cartdb.objects.filter(commande=commande)
                    for item in items:
                        if item.qte > item.produit.stock:
                            nb_produit += 1
                            produit_list += "<li>" + str(item.produit.nom) + " (Manque : " + str(item.qte - item.produit.stock) + ")</li>"

            else:
                if nb_produit > 0:
                    message = format_html("Impossible de terminer toutes les commandes.<br><strong>" + str(nb_produit) + "</strong> produits ne disposent pas d'un stock final suffisant !<ul>"+produit_list+"</ul>")
                    messages.error(request, message)
                else:
                    for commande in commandes:
                        old_statut = commande.statut.nom
                        if commande.statut == wait_statut:
                            commande.statut = cancel_statut
                        else:
                            commande.statut = done_statut
                        commande.save()
                        log_order("Commande", str(request.user), commande.pk, "End", "statut", old_statut, commande.statut.nom)

                        if commande.statut.nom == "En cours" or commande.statut.nom == "Validée":
                            items = Cartdb.objects.filter(commande=commande)
                            for item in items:
                                old_final = item.produit.stock
                                new_final = item.produit.stock - item.qte

                                item.produit.stock -= item.qte
                                item.produit.save()

                                log_produit("Produit", str(request.user), item.produit.pk, commande.pk, "Update", "sf", old_final, new_final)

                    message = format_html("Les commandes ont bien toutes étés <b>Terminées</b> pour la période selectionnée !<br>Ne sont concernées que les commandes [<b>Validées</b>] ou [<b>En cours</b>] !")
                    messages.success(request, message)

            return redirect('order:order-administration')

    context = {
        'inventaires': inventaires,
        'commandes': commandes,
        'form': form,
        'formAction': formAction,
        'previous_page': previous_page,
    }
    return render(request, "order/form_reset_order.html", context)


def warning_order(request):
    inventaire = Inventaire.objects.get(actif=True)
    statut_valide = Statut.objects.get(nom="Validée")
    statut_encours = Statut.objects.get(nom="En cours")
    statut_attente = Statut.objects.get(nom="En attente")
    statut_future = Statut.objects.get(nom="Pré-commande")
    statut_annulee = Statut.objects.get(nom="Annulée")
    statut_terminee = Statut.objects.get(nom="Terminée")

    if request.user.is_staff:
        produits_commande = {}
        form = SearchOrderForm()
        queryset = Commande.objects\
            .filter(inventaire=inventaire, statut__in=[statut_valide, statut_encours])\
            .exclude(statut__in=[statut_attente, statut_annulee, statut_terminee, statut_future])\
            .order_by('-id')
            # .distinct()
        # print(queryset.query)
        # queryset = list(set(queryset))
        orders_warning = []
        orders = queryset
        produits_warning = []
        if len(queryset) > 0:
            for commande in queryset:
                produits = Cartdb.objects.filter(commande=commande)
                list_produit = []
                for produit in produits:
                    if total_qte_inventaire_progress(produit.produit) > produit.produit.stock:
                        orders_warning.append(commande)
                        produits_warning.append(produit.produit)
                    list_produit.append(produit)
                produits_commande[commande.id] = list_produit

            orders_warning = list(set(orders_warning))
            orders_warning.sort(key=lambda x: -x.id)

        title = "Commandes"
        header = "Ajouter un Produit"
        javascript = "Cela va supprimer la commande"
        formAction = "order:manage-order"
        previous_page = reverse('order:order-administration')

        context_header = {
            'header': header,
            'javascript': javascript,
        }
        context = {
            'orders_warning': orders_warning,
            'produits_commande': produits_commande,
            'produits_warning': produits_warning,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'formAction': formAction,
            'form': form,
            'orders_list': queryset,
        }
        return render(request, "order/manage_warning_order.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('onlineshop:produit-list')
