from datetime import datetime
from sqlite3 import DatabaseError
from tablib import Dataset
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib import messages
from django.utils.html import format_html
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.core.files.storage import FileSystemStorage
from django.db.models import Sum
from django.db.models import Q
from django.utils.text import slugify
from django.http import JsonResponse
from django.forms import formset_factory

from account.decorators import unauthenticated_user
from account.decorators import allowed_users

from onlineshop.models import *
from onlineshop.resources import *
from onlineshop.forms import *
from onlineshop.core import *
from cart.forms import CartAddProduitForm
from pepiniere import settings
from order.models import *
from order.core import *
from .forms import *
from order.forms import *
import xlsxwriter
import io
import json
import pandas as pd

from openpyxl import load_workbook


@login_required
@staff_member_required
def full_admin(request):
    if request.user.is_authenticated:
        admin = AccessMode.objects.filter(user=request.user).first()
        if admin:
            if admin.admin is True:
                admin.admin = False
            else:
                admin.admin = True
            admin.save()
        else:
            AccessMode.objects.create(user=request.user, admin=True)

    return redirect('produit-list')


def create_tab_dict(items, menu):
    itemslist = []
    for item in items:
        if menu == "variete":
            if item.get_variete() not in itemslist:
                itemslist.append(item.get_variete())
        if menu == "portegreffe":
            if item.get_portegreffe() not in itemslist:
                itemslist.append(item.get_portegreffe())
        if menu == "spec":
            if item.get_spec() not in itemslist:
                itemslist.append(item.get_spec())
    return itemslist


def produit_list(request):
    formAction = 'onlineshop:produit-list'

    form = SearchProduitForm(request.GET or None)

    if 'p' in request.GET:
        order_value = request.GET['p']
        if order_value == 'total':
            queryset = Produit.objects.annotate(sum=Sum('Cartdbs__qte')).order_by('-sum')
        else:
            queryset = Produit.objects.all().order_by(order_value)
        request.session['p'] = request.GET['p']
    elif 'p' in request.session:
        order_value = request.session['p']
        if order_value == 'total':
            queryset = Produit.objects.annotate(sum=Sum('Cartdbs__qte')).order_by('-sum')
        else:
            queryset = Produit.objects.all().order_by(order_value)
    else:
        queryset = Produit.objects.all().order_by('variete', 'espece', 'portegreffe')

    if request.method == 'GET':
        if form.is_valid():
            espece = form.cleaned_data['especes']
            variete = form.cleaned_data['varietes']
            portegreffe = form.cleaned_data['portegreffes']
            spec = form.cleaned_data['specs']
            stock = form.cleaned_data['stock']
            gaf = form.cleaned_data['gaf']

            if not espece is None:
                queryset = queryset.filter(espece=espece)
            if not variete is None:
                queryset = queryset.filter(variete=variete)
            if not portegreffe is None:
                queryset = queryset.filter(portegreffe=portegreffe)
            if not spec is None:
                queryset = queryset.filter(spec=spec)
            if stock is True:
                queryset = queryset.filter(stock_bis__gt=0)
            if gaf is True:
                queryset = queryset.filter(gaf=True)
        else:
            queryset = queryset.filter(stock_bis__gt=0)
    else:
        queryset = queryset.filter(stock_bis__gt=0)

    if request.user.is_staff is False:
        admin_mode = False
        queryset = queryset.filter(gaf=False, stock_bis__gt=0)
    else:
        try:
            admin_mode = AccessMode.objects.get(user=request.user).admin
        except:
            admin_mode = False

    paginator = Paginator(queryset, 50)
    get_data = request.GET.copy()
    page = get_data.pop('page', None)
    p = get_data.pop('p', None)
    if 'page' in request.GET:
        page = request.GET['page']
        request.session['page_p'] = page
    elif 'page_p' in request.session:
        page = request.session['page_p']

    try:
        produits = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        produits = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        produits = paginator.page(paginator.num_pages)

    context = {'produits': produits,
               'produits_list': queryset,
               'paginate': True,
               'form': form,
               'formAction': formAction,
               'admin_mode': admin_mode,
               'query_string': get_data.urlencode(),
               }
    return render(request, 'onlineshop/list.html', context)


def produit_detail(request, id):
    produit = get_object_or_404(Produit, id=id, available=True)
    cart_produit_form = CartAddProduitForm
    context = {
        'produit': produit,
        'cart_produit_form': cart_produit_form,
        'previous_page': 'onlineshop:produit-list',
    }
    return render(request, 'onlineshop/detail.html', context)


# ****************************************************************************************************************
# ADMINISTRATION
# ****************************************************************************************************************
@login_required
@staff_member_required
def onlineshop_administration(request):
    produits = Produit.objects.all()
    form = FormProduitList(request.POST or None)
    context = {'produits':produits, 'form': form}
    if request.user.is_staff:
        return render(request, 'onlineshop/administration_menu_onlineshop.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# PRODUITS *****************************************************************************************************
@login_required
@staff_member_required
def manage_produit(request):
    form = SearchProduitForm(request.GET or None)

    if request.user.is_staff:

        if 'p' in request.GET:
            order_value = request.GET['p']
            if order_value == 'total':
                queryset = Produit.objects.annotate(sum=Sum('Cartdbs__qte')).order_by('-sum')
            else:
                queryset = Produit.objects.all().order_by(order_value)
            request.session['p'] = request.GET['p']
        elif 'p' in request.session:
            order_value = request.session['p']
            if order_value == 'total':
                queryset = Produit.objects.annotate(sum=Sum('Cartdbs__qte')).order_by('-sum')
            else:
                queryset = Produit.objects.all().order_by(order_value)
        else:
            queryset = Produit.objects.all().order_by('variete', 'espece', 'portegreffe')

        if request.method == 'GET':
            if form.is_valid():
                espece = form.cleaned_data['especes']
                variete = form.cleaned_data['varietes']
                portegreffe = form.cleaned_data['portegreffes']
                spec = form.cleaned_data['specs']
                stock = form.cleaned_data['stock']
                gaf = form.cleaned_data['gaf']

                if not espece is None:
                    queryset = queryset.filter(espece=espece)
                if not variete is None:
                    queryset = queryset.filter(variete=variete)
                if not portegreffe is None:
                    queryset = queryset.filter(portegreffe=portegreffe)
                if not spec is None:
                    queryset = queryset.filter(spec=spec)
                if stock is True:
                    queryset = queryset.filter(stock_bis__gt=0)
                if gaf is True:
                    queryset = queryset.filter(gaf=True)
            else:
                queryset = queryset.filter(stock_bis__gt=0)
        title = "Produits"
        header = "Ajouter un Produit"
        javascript = "Cela va supprimer le produit"
        formAction = "onlineshop:manage-produit"
        previous_page = reverse('onlineshop:onlineshop-administration')

        paginator = Paginator(queryset, 50)
        get_data = request.GET.copy()
        page = get_data.pop('page', None)
        p = get_data.pop('p', None)
        if 'page' in request.GET:
            page = request.GET['page']
            request.session['page_p'] = page
        elif 'page_p' in request.session:
            page = request.session['page_p']

        try:
            produits = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            produits = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            produits = paginator.page(paginator.num_pages)

        context_header = {
            'header': header,
            'javascript': javascript,
        }

        GET_params = request.GET.copy()

        context = {
            'produits': produits,
            'produits_list': queryset,
            'paginate': True,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'formAction': formAction,
            'form': form,
            'query_string': get_data.urlencode(),
            'GET_params': GET_params,
        }

        return render(request, 'onlineshop/manage_produit.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('onlineshop:produit-list')


@login_required
@staff_member_required
def add_produit(request):
    if request.user.is_staff:
        title = "Espèces"
        form = FormProduit(request.POST or None)
        message = "Produit ajouté avec succès !"

        previous_page = reverse('onlineshop:manage-produit')
        formAction = 'onlineshop:add-produit'

        if request.POST:
            if form.is_valid():
                instance = form.instance
                stock = instance.stock
                stock_bis = instance.stock_bis
                stock_future = instance.stock_future

                if stock is None or stock < 0 or not str(stock).isnumeric():
                    instance.stock = 0
                if stock_bis is None or stock_bis < 0 or not str(stock_bis).isnumeric():
                    instance.stock_bis = 0
                if stock_future is None or stock_future < 0 or not str(stock_future).isnumeric():
                    instance.stock_future = 0
                if instance.stock_bis > instance.stock:
                    instance.stock = instance.stock_bis

                if instance.nom is None or instance.nom == "":
                    espece = str(instance.espece)
                    variete = str(instance.variete)
                    portegreffe = str(instance.portegreffe)
                    instance.nom = espece + "-" + variete + "-" + portegreffe

                if check_produit_exist(instance.espece, instance.variete, instance.portegreffe, instance.spec) > 0:
                    messages.error(request, "Ce produit existe déjà ! Merci de réessayer.")
                    return redirect('onlineshop:manage-produit')

                instance.save()

                log_produit(str(request.user), instance.pk, None, 'New', 'produit', '', instance.nom)
                log_produit(str(request.user), instance.pk, None, 'New', 'sf', '', instance.stock)
                log_produit(str(request.user), instance.pk, None, 'New', 'sb', '', instance.stock_bis)
                log_produit(str(request.user), instance.pk, None, 'New', 'sp', '', instance.stock_future)
                log_produit(str(request.user), instance.pk, None, 'New', 'prix', 0, float(instance.prix))
                log_produit(str(request.user), instance.pk, None, 'New', 'description', '', instance.description)
                log_produit(str(request.user), instance.pk, None, 'New', 'available', '', instance.available)
                log_produit(str(request.user), instance.pk, None, 'New', 'gaf', '', instance.gaf)
                messages.success(request, message)
                return redirect('onlineshop:manage-produit')
        context_header = {
        }
        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "onlineshop/form_produit.html", context)
    else:
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_produit(request, produit_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "Produits"
        produit = Produit.objects.get(id=produit_id)

        old_nom = produit.nom
        old_prix = produit.prix
        old_final = produit.stock
        old_bis = produit.stock_bis
        old_future = produit.stock_future
        old_description = produit.description
        old_available = produit.available
        old_gaf = produit.gaf

        form = FormProduit(request.POST or None, instance=produit)
        message = "Produit modifié avec succès !"

        previous_page = reverse('onlineshop:manage-produit')
        formAction = 'onlineshop:edit-produit', produit_id

        if request.POST:
            if form.is_valid():
                instance = form.instance
                stock = instance.stock
                stock_bis = instance.stock_bis
                stock_future = instance.stock_future
                if stock is None or stock < 0 or not str(stock).isnumeric():
                    instance.stock = 0
                if stock_bis is None or stock_bis < 0 or not str(stock_bis).isnumeric():
                    instance.stock_bis = 0
                if stock_future is None or stock_future < 0 or not str(stock_future).isnumeric():
                    instance.stock_future = 0
                if instance.stock_bis > instance.stock:
                    instance.stock = instance.stock_bis

                if instance.nom is None or instance.nom == "":
                    espece = str(instance.espece)
                    variete = str(instance.variete)
                    portegreffe = str(instance.portegreffe)
                    instance.nom = espece + "-" + variete + "-" + portegreffe

                obj = instance.save()

                new_nom = instance.nom
                new_prix = instance.prix
                new_final = instance.stock
                new_bis = instance.stock_bis
                new_future = instance.stock_future
                new_description = instance.description
                new_available = instance.available
                new_gaf = instance.gaf
                if old_nom != new_nom:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'produit', old_nom, new_nom)
                if old_prix != new_prix:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'prix', old_prix, new_prix)
                if old_final != new_final:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'sf', old_final, new_final)
                if old_bis != new_bis:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'sb', old_bis, new_bis)
                if old_future != new_future:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'sp', old_future, new_future)
                if old_description != new_description:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'description', old_description, new_description)
                if old_available != new_available:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'available', old_available, new_available)
                if old_gaf != new_gaf:
                    log_produit(str(request.user), instance.pk, None, 'Edit', 'gaf', old_gaf, new_gaf)
                messages.success(request, message)
                return redirect('onlineshop:manage-produit')
        context = {
            'produit_id': produit_id,
            'produit': produit,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "onlineshop/form_produit.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_produit(request, produit_id):
    if request.user.is_staff:
        title = "Produits"
        produit = Produit.objects.get(id=produit_id)

        old_nom = produit.nom
        old_prix = produit.prix
        old_final = produit.stock
        old_bis = produit.stock_bis
        old_future = produit.stock_future
        old_description = produit.description
        old_available = produit.available
        old_gaf = produit.gaf

        in_order = Cartdb.objects.filter(produit=produit)
        if len(in_order) > 0:
            message = "Impossible de supprimer ce produit ... présent dans une ou plusieurs commande !"
            messages.error(request, message)
            return redirect('onlineshop:manage-produit')

        produit.delete()
        log_produit(str(request.user), produit.pk, None, 'Delete', 'produit', old_nom, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'prix', old_prix, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'sf', old_final, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'sb', old_bis, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'sp', old_future, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'description', old_description, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'available', old_available, '')
        log_produit(str(request.user), produit.pk, None, 'Delete', 'gaf', old_gaf, '')

        message = "Produit supprimé avec succès !"
        messages.success(request, message)
        return redirect('onlineshop:manage-produit')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def manage_greffons(request):
    form = SearchGreffonsForm(request.GET or None)
    couleurs = Couleur.objects.all()
    inventaires = Inventaire.objects.all()

    if request.user.is_staff:
        if 'g' in request.GET:
            order_value = request.GET['g']
            queryset = Greffons.objects.all().order_by(order_value)
            request.session['g'] = request.GET['g']
        elif 'g' in request.session:
            order_value = request.session['g']
            queryset = Greffons.objects.all().order_by(order_value)
        else:
            queryset = Greffons.objects.all().order_by('produit')

        if request.method == 'GET':
            print(form.errors)
            if form.is_valid():
                espece = form.cleaned_data['especes']
                variete = form.cleaned_data['varietes']
                portegreffe = form.cleaned_data['portegreffes']
                spec = form.cleaned_data['specs']
                start_date = form.cleaned_data['start_date']
                end_date = form.cleaned_data['end_date']
                couleur = form.cleaned_data['couleur']
                inventaire = form.cleaned_data['inventaire']

                if espece is not None:
                    queryset = queryset.filter(produit__espece=espece)
                if variete is not None:
                    queryset = queryset.filter(produit__variete=variete)
                if portegreffe is not None:
                    queryset = queryset.filter(produit__portegreffe=portegreffe)
                if spec is not None:
                    queryset = queryset.filter(produit__spec=spec)
                if start_date is not None and start_date != "":
                    queryset = queryset.filter(date__gte=start_date)
                if end_date is not None and end_date != "":
                    queryset = queryset.filter(date__lte=end_date)
                if couleur.exists():
                    queryset = queryset.filter(couleur__in=couleur)
                if inventaire is not None:
                    queryset = queryset.filter(inventaire=inventaire)
                else:
                    inventaire_actif = Inventaire.objects.get(actif=True)
                    queryset = queryset.filter(inventaire=inventaire_actif)
            else:
                inventaire_actif = Inventaire.objects.get(actif=True)
                queryset = queryset.filter(inventaire=inventaire_actif)

        title = "Greffons"
        header = "Ajouter un Greffon"
        javascript = "Cela va supprimer le Greffon"
        formAction = "onlineshop:manage-greffons"
        previous_page = reverse('onlineshop:onlineshop-administration')

        paginator = Paginator(queryset, 50)
        get_data = request.GET.copy()
        page = get_data.pop('page', None)
        g = get_data.pop('g', None)
        if 'page' in request.GET:
            page = request.GET['page']
            request.session['page_g'] = page
        elif 'page_g' in request.session:
            page = request.session['page_g']
        try:
            greffons = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            greffons = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            greffons = paginator.page(paginator.num_pages)

        context_header = {
            'header': header,
            'javascript': javascript,
        }

        GET_params = request.GET.copy()

        context = {
            'greffons': greffons,
            'greffons_list': queryset,
            'paginate': True,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'formAction': formAction,
            'form': form,
            'query_string': get_data.urlencode(),
            'couleurs': couleurs,
            'inventaires': inventaires,
            'GET_params': GET_params,
        }

        return render(request, 'onlineshop/manage_greffons.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('onlineshop:produit-list')


@login_required
@staff_member_required
def valid_greffons(request):
    if request.POST:
        try:
            inventaire = Inventaire.objects.get(pk=request.POST.get('inventaire'))
        except:
            messages.error(request, "Période inexistante !")
            return redirect('onlineshop:manage-greffons')

        # check_stock_produits = Produit.objects.filter(Q(stock__gt=0) | Q(stock_bis__gt=0))
        # if len(check_stock_produits) > 0:
        #     messages.error(request, "Attention certains produits ont encore du stock ... Réinitaliser les stocks avant de poursuivre !")
        #    return redirect('onlineshop:manage-greffons')

        greffons = Greffons.objects.filter(inventaire=inventaire)
        # Pour Chaque Greffon
        # JUIN : REUSSI = REALISE > MAJ DES STOCK ET STOCK BIS
        # AVRIL : REALISE > MAJ DES STOCKS ET STOCKS BIS
        for greffon in greffons:
            if request.POST.get('stock') == "realises":
                greffon.produit.stock = greffon.realise
                greffon.produit.stock_bis = greffon.realise
            elif request.POST.get('stock') == "reussis":

                if greffon.produit.stock == greffon.produit.stock_bis:
                    greffon.produit.stock_bis = greffon.reussi
                else:
                    produits_commandes = greffon.produit.stock_bis - greffon.produit.stock  # Nombre negatif
                    test_en_cours = greffon.reussi + produits_commandes
                    if test_en_cours < 0:
                        print("Stock en cours negatif", test_en_cours)
                        greffon.produit.stock_bis = 0
                    else:
                        print("On est encore pas mal", test_en_cours)
                        greffon.produit.stock_bis = test_en_cours
                greffon.produit.stock = greffon.reussi
            else:
                if greffon.reussi > 0:
                    if greffon.produit.stock == greffon.produit.stock_bis:
                        greffon.produit.stock_bis = greffon.reussi
                    greffon.produit.stock = greffon.reussi
                else:
                    greffon.produit.stock = greffon.realise
                    greffon.produit.stock_bis = greffon.realise
            greffon.produit.save()
            messages.success(request, "Stock mis à jour avec succès !")

    return redirect('onlineshop:manage-greffons')


@login_required
@staff_member_required
def init_greffons(request):
    if request.user.is_staff:
        if request.POST:
            inventaire_id = request.POST.get('inventaire')
            inventaire = Inventaire.objects.get(pk=inventaire_id)

            # TEST SI INITIALISATION DEJA REALISE SUR LA PERIODE
            check_inventaire_greffons = Greffons.objects.filter(inventaire=inventaire)
            if len(check_inventaire_greffons) > 0:
                message = "Greffons déjà initialisés sur cette période !"
                messages.error(request, message)
                return redirect('onlineshop:manage-greffons')

            produits = Produit.objects.all()
            for produit in produits:
                greffon = Greffons.objects.create(produit=produit, date=datetime.now(), inventaire=inventaire, comm=produit.stock_future)
                greffon.save()
            message = "Greffons initialisés avec succès pour cette nouvelle période !"
            messages.success(request, message)
            return redirect('onlineshop:manage-greffons')
    return redirect('onlineshop:manage-greffons')

@login_required
@staff_member_required
def add_greffon(request):
    if request.user.is_staff:
        title = "Greffons"
        inventaire = Inventaire.objects.get(actif=True)
        greffons = Greffons.objects.filter(inventaire=inventaire)
        produit_restant = Produit.objects.exclude(Greffons__in=greffons)
        form = FormGreffon(request.POST or None, produit_restant, initial={'inventaire': inventaire})

        previous_page = reverse('onlineshop:manage-greffons')
        formAction = 'onlineshop:add-greffon'

        if request.POST:
            if form.is_valid():
                instance = form.instance
                produit = instance.produit

                test_greffons = Greffons.objects.filter(produit=produit, inventaire=instance.inventaire)
                if len(test_greffons) > 0:
                    message = "Greffon déjà présent pour cette période !"
                    messages.error(request, message)
                    return redirect('onlineshop:manage-greffons')

                greffons = instance.greffons
                comm = instance.comm
                objectif = instance.objectif
                realise = instance.realise
                reussi = instance.reussi

                if greffons is None or greffons == "":
                    instance.greffons = 0

                if comm is None or comm == "":
                    instance.comm = 0

                if objectif is None or objectif == "":
                    instance.objectif = 0

                if realise is None or realise == "":
                        instance.realise = 0

                if reussi is None or reussi == "":
                    instance.reussi = 0

                if instance.date is None:
                    instance.date = datetime.now()

                if instance.inventaire is None:
                    instance.inventaire = inventaire

                instance.save()
                message = "Greffon ajouté avec succès !"
                messages.success(request, message)
                return redirect('onlineshop:manage-greffons')
        context_header = {
        }
        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "onlineshop/form_greffon.html", context)
    else:
        return redirect('produit-list')


@login_required
@staff_member_required
def edit_greffon(request, greffon_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "Greffons"
        greffon = Greffons.objects.get(id=greffon_id)

        form = FormEditGreffon(request.POST or None, instance=greffon, initial={'produit': greffon.produit})
        message = "Greffon modifié avec succès !"

        previous_page = reverse('onlineshop:manage-greffons')
        formAction = 'onlineshop:edit-greffon', greffon_id

        if request.POST:
            if form.is_valid():
                instance = form.instance
                # print(instance.produit)
                # instance.produit = greffon.produit
                qte = instance.greffons
                comm = instance.comm
                realise = instance.realise
                objectif = instance.objectif
                reussi = instance.reussi
                rang = instance.rang
                couleur = instance.couleur
                date = instance.date

                instance.qte = 0 if qte is None or qte == "" else qte
                instance.comm = 0 if comm is None or comm == "" else comm
                instance.objectif = 0 if objectif is None or objectif == "" else objectif
                instance.realise = 0 if realise is None or realise == "" else realise
                instance.reussi = 0 if reussi is None or reussi == "" else reussi
                instance.date = date
                instance.couleur = couleur
                instance.rang = rang
                obj = instance.save()

                messages.success(request, message)
                return redirect('onlineshop:manage-greffons')
        context = {
            'greffon_id': greffon_id,
            'greffon': greffon,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "onlineshop/form_greffon.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_greffon(request, greffon_id):
    if request.user.is_staff:
        title = "Greffons"
        greffon = Greffons.objects.get(id=greffon_id)
        form = FormGreffon(request.POST or None, instance=greffon)
        message = "Greffon supprimé avec succès !"

        greffon.delete()
        messages.success(request, message)
        return redirect('onlineshop:manage-greffons')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# MISE A JOUR DES STOCKS D'UN PRODUIT
@login_required
@staff_member_required
def edit_qte_greffon(request):
    if request.method == 'POST' and request.is_ajax:
        data_string = request.POST.get('json_data')
        data_dict = json.loads(data_string)
        try:
            greffon_id = data_dict['code']
            qte = data_dict['greffon']
            comm = data_dict['comm']
            objectif = data_dict['objectif']
            realise = data_dict['realise']
            reussi = data_dict['reussi']
            date = datetime.strptime(data_dict['date'], '%d/%m/%Y')
            couleur_id = data_dict['couleur']
            inventaire_id = data_dict['inventaire']
        except Exception as e:
            print("Exception", e)

        if couleur_id is None or couleur_id == "":
            couleur = None
        else:
            couleur = Couleur.objects.get(id=couleur_id)

        if inventaire_id is None or inventaire_id == "":
            inventaire = None
        else:
            inventaire = Inventaire.objects.get(id=inventaire_id)

        try:
            greffon = Greffons.objects.get(id=greffon_id)
        except:
            messages.error(request, "Le greffon n'existe pas !")
            return redirect('onlineshop:manage-greffons')

        try:
            greffon.greffons = int(qte)
            greffon.comm = int(comm)
            greffon.objectif = int(objectif)
            greffon.realise = int(realise)
            greffon.reussi = int(reussi)
            greffon.date = date
            greffon.couleur = couleur
            greffon.inventaire = inventaire
            greffon.save()
        except Exception as e:
            print(e)

        messages.success(request, "Infos mise à jour pour le greffon ;)")
    return redirect('onlineshop:manage-greffons')


@login_required
@staff_member_required
def manage_couleurs(request):
    couleurs = Couleur.objects.all()

    previous_page = reverse('onlineshop:onlineshop-administration')
    context = {
        'couleurs': couleurs,
        'title': 'Gestion des couleurs',
        'previous_page': previous_page,
    }
    return render(request, 'onlineshop/manage_couleurs.html', context)


# *************************************************************************************
# TVA (ADD/UPDATE/DELETE)
# *************************************************************************************
@login_required
@staff_member_required
def add_couleur(request):
    if request.user.is_staff:
        title = "AJOUTER UNE COULEUR"
        form = FormCouleur(request.POST or None)
        previous_page = reverse('onlineshop:manage-greffons')
        formAction = 'onlineshop:add-couleur'

        if request.method == 'POST':
            if form.is_valid():
                nom = form.cleaned_data['nom']
                couleur = form.cleaned_data['couleur']
                form.save()

                message = "Nouvelle couleur ajoutée avec succès !"
                messages.success(request, message)
            else:
                message = "Une erreur s'est produite"
                messages.error(request, message)
            return redirect('onlineshop:manage-couleurs')

        context_header = {
        }

        context = {
            'form': form,
            'formAction': formAction,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "onlineshop/form_couleur.html", context)
    else:
        return redirect('onlineshop:manage-couleurs')


@login_required
@staff_member_required
def edit_couleur(request, couleur_id):
    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        title = "COULEUR"
        couleur = Couleur.objects.get(id=couleur_id)
        form = FormCouleur(request.POST or None, instance=couleur)

        previous_page = reverse('onlineshop:manage-couleurs')
        formAction = 'onlineshop:edit-couleur', couleur_id

        if request.POST:
            if form.is_valid():
                nom = form.cleaned_data['nom']
                couleur = form.cleaned_data['couleur']

                message = "Couleur moidifiée avec succès !"
                messages.success(request, message)
                instance = form.instance
                obj = instance.save()

                return redirect('onlineshop:manage-couleurs')
        context = {
            'couleur_id': couleur_id,
            'couleur': couleur,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'title': title,
        }
        return render(request, "onlineshop/form_couleur.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_couleur(request, couleur_id):
    if request.user.is_staff:
        try:
            couleur = Couleur.objects.get(id=couleur_id)
            couleur.delete()
            message = "Couleur supprimée avec succès !"
            messages.success(request, message)
            return redirect('onlineshop:manage-couleurs')
        except:
            message = "Couleur inexistante !"
            messages.error(request, message)
            return redirect('onlineshop:manage-couleurs')
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# ESPECES / VARIETES / PORTE GREFFE / SPECS ************************************************************************************
@login_required
@staff_member_required
def manage_data(request, categorie):
    categories = ['espece', 'variete', 'portegreffe', 'spec']
    if categorie not in categories:
        messages.error(request, "La catégorie spécifiée n'existe pas !")
        return redirect('onlineshop:onlineshop-administration')

    if request.user.is_staff:
        if categorie == 'espece':
            title = "Espèces"
            datas = Espece.objects.all()
            header = "Ajouter une Espèce"
            javascript = "Cela va supprimer l'espèce"

        if categorie == 'variete':
            title = "Variétés"
            datas = Variete.objects.all()
            header = "Ajouter une Variété"
            javascript = "Cela va supprimer la variété"

        if categorie == 'portegreffe':
            title = "Porte-Greffe"
            datas = PorteGreffe.objects.all()
            header = "Ajouter un Porte-Greffe"
            javascript = "Cela va supprimer le porte-greffe"

        if categorie == 'spec':
            title = "Spécialités"
            datas = Spec.objects.all()
            header = "Ajouter une Spécialité"
            javascript = "Cela va supprimer la spécialité"

        previous_page = reverse('onlineshop:onlineshop-administration')

        context_header = {
            'header': header,
            'javascript': javascript,
        }
        context = {
            'categorie': categorie,
            'datas': datas,
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
        }

        return render(request, 'onlineshop/manage_data.html', context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def add_data(request, categorie):
    categories = ['espece', 'variete', 'portegreffe', 'spec']
    if categorie not in categories:
        messages.error(request, "La catégorie spécifiée n'existe pas !")
        return redirect('onlineshop:onlineshop-administration')

    if request.user.is_staff:
        if categorie == 'espece':
            title = "Espèces"
            form = FormEspece(request.POST or None)
            message = "Espèce ajoutée avec succès !"

        if categorie == 'variete':
            title = "Variétés"
            form = FormVariete(request.POST or None)
            message = "Variété ajoutée avec succès !"

        if categorie == 'portegreffe':
            title = "Porte-Greffes"
            form = FormPorteGreffe(request.POST or None)
            message = "Porte-Greffe ajouté avec succès !"

        if categorie == 'spec':
            title = "Spécialités"
            form = FormSpec(request.POST or None)
            message = "Spécialité ajoutée avec succès !"

        previous_page = reverse('onlineshop:manage-data', kwargs={'categorie': categorie})
        formAction = 'onlineshop:add-data', categorie

        if request.POST:
            if form.is_valid():
                obj = form.save()
                messages.success(request, message)
                return redirect('onlineshop:manage-data', categorie)
        context_header = {
        }
        context = {
            'form': form,
            'previous_page': previous_page,
            'formAction': formAction,
            'categorie': categorie,
            'title': title,
            'context_header': context_header,
        }
        return render(request, "onlineshop/form_data.html", context)
    else:

        return redirect('produit-list')


@login_required
@staff_member_required
def edit_data(request, categorie, data_id):
    # TEST SI CATEGORIE EXISTE
    categories = ['espece', 'variete', 'portegreffe', 'spec']
    if categorie not in categories:
        messages.error(request, "La catégorie spécifiée n'existe pas !")
        return redirect('onlineshop:onlineshop-administration')

    # RECUPERATION DES INFOS SELON CATEGORIE SI ADMIN CONNECTE
    if request.user.is_staff:
        if categorie == 'espece':
            title = "Espèces"
            data = Espece.objects.get(id=data_id)
            form = FormEspece(request.POST or None, instance=data)
            message = "Espèce modifiée avec succès !"

        if categorie == 'variete':
            title = "Variétés"
            data = Variete.objects.get(id=data_id)
            form = FormVariete(request.POST or None, instance=data)
            message = "Variété modifiée avec succès !"

        if categorie == 'portegreffe':
            title = "Porte-Greffes"
            data = PorteGreffe.objects.get(id=data_id)
            form = FormPorteGreffe(request.POST or None, instance=data)
            message = "Porte-Greffe modifiée avec succès !"

        if categorie == 'spec':
            title = "Spécialités"
            data = Spec.objects.get(id=data_id)
            form = FormSpec(request.POST or None, instance=data)
            message = "Spécialité modifiée avec succès !"

        previous_page = reverse('onlineshop:manage-data', kwargs={'categorie': categorie})
        formAction = 'onlineshop:edit-data', categorie, data_id

        if request.POST:
            if form.is_valid():
                obj = form.save()
                messages.success(request, message)
                return redirect('onlineshop:manage-data', categorie)
        context = {
            'data_id': data_id,
            'data': data,
            'formAction': formAction,
            'form': form,
            'previous_page': previous_page,
            'categorie': categorie,
            'title': title,
        }
        return render(request, "onlineshop/form_data.html", context)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


@login_required
@staff_member_required
def delete_data(request, categorie, data_id):
    categories = ['espece', 'variete', 'portegreffe', 'spec']
    if categorie not in categories:
        messages.error(request, "La catégorie spécifiée n'existe pas !")
        return redirect('onlineshop:onlineshop-administration')

    if request.user.is_staff:
        if categorie == 'espece':
            data = Espece.objects.get(id=data_id)
            if len(Produit.objects.filter(espece=data)) > 0:
                message = "Impossible de supprimer l'espèce, un ou plusieurs produits sont encore associés à cette espèce !"
                messages.error(request, message)
                return redirect('onlineshop:manage-data', categorie)
            message = "Espèce supprimée avec succès !"
        if categorie == 'variete':
            data = Variete.objects.get(id=data_id)
            if len(Produit.objects.filter(variete=data)) > 0:
                message = "Impossible de supprimer la variété, un ou plusieurs produits sont encore associés à cette variété !"
                messages.error(request, message)
                return redirect('onlineshop:manage-data', categorie)
            message = "Variété supprimée avec succès !"
        if categorie == 'portegreffe':
            data = PorteGreffe.objects.get(id=data_id)
            if len(Produit.objects.filter(portegreffe=data)) > 0:
                message = "Impossible de supprimer le porte-greffe, un ou plusieurs produits sont encore associés à ce porte-greffe !"
                messages.error(request, message)
                return redirect('onlineshop:manage-data', categorie)
            message = "Porte-Greffe supprimé avec succès !"
        if categorie == 'spec':
            data = Spec.objects.get(id=data_id)
            if len(Produit.objects.filter(spec=data)) > 0:
                message = "Impossible de supprimer la spécialité, un ou plusieurs produits sont encore associés à cete spécialité !"
                messages.error(request, message)
                return redirect('onlineshop:manage-data', categorie)
            message = "Spécialité supprimée avec succès !"

        data.delete()
        messages.success(request, message)
        return redirect('onlineshop:manage-data', categorie)
    else:
        messages.error(request, "Vous n'avez pas les droits !")
        return redirect('produit-list')


# ----------------------------------------------------------------------------------------------------------------------------------------
# MANAGE GLOBAL
# ----------------------------------------------------------------------------------------------------------------------------------------
@login_required
@staff_member_required
def export_produits_xls(request):
    output = io.BytesIO()

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output)
    worksheet_produits = workbook.add_worksheet("PRODUITS")
    worksheet_especes = workbook.add_worksheet("ESPECES")
    worksheet_varietes = workbook.add_worksheet("VARIETES")
    worksheet_portegreffes = workbook.add_worksheet("PORTE-GREFFES")
    worksheet_specialites = workbook.add_worksheet("SPECIALITES")

    # PRODUITS
    worksheet_produits.write(0, 0, 'ID')
    worksheet_produits.write(0, 1, 'NOM')
    worksheet_produits.write(0, 2, 'SLUG')
    worksheet_produits.write(0, 3, 'DESCRIPTION')
    worksheet_produits.write(0, 4, 'PRIX')
    worksheet_produits.write(0, 5, 'STOCK')
    worksheet_produits.write(0, 6, 'STOCK BIS')
    worksheet_produits.write(0, 7, 'AVAILABLE')
    worksheet_produits.write(0, 8, 'ESPECE')
    worksheet_produits.write(0, 9, 'VARIETE')
    worksheet_produits.write(0, 10, 'PORTEGREFFE')
    worksheet_produits.write(0, 11, 'SPEC')
    worksheet_produits.write(0, 12, 'GAF')

    produits = Produit.objects.all()
    row = 1
    for produit in produits:
        worksheet_produits.write(row, 0, produit.id)
        worksheet_produits.write(row, 1, produit.nom)
        worksheet_produits.write(row, 2, produit.slug)
        worksheet_produits.write(row, 3, produit.description)
        worksheet_produits.write(row, 4, produit.prix)
        worksheet_produits.write(row, 5, produit.stock)
        worksheet_produits.write(row, 6, produit.stock_bis)
        worksheet_produits.write(row, 7, produit.available)
        worksheet_produits.write(row, 8, produit.espece.id)
        worksheet_produits.write(row, 9, produit.variete.id)
        worksheet_produits.write(row, 10, produit.portegreffe.id)
        if not produit.spec is None:
            worksheet_produits.write(row, 11, produit.spec.id)
        worksheet_produits.write(row, 12, produit.gaf)

        row += 1

    # ESPECES
    worksheet_especes.write(0, 0, 'ID')
    worksheet_especes.write(0, 1, 'NOM')
    worksheet_especes.write(0, 2, 'SLUG')

    produits = Espece.objects.all()
    row = 1
    for produit in produits:
        worksheet_especes.write(row, 0, produit.id)
        worksheet_especes.write(row, 1, produit.nom)
        worksheet_especes.write(row, 2, produit.slug)
        row += 1

    # VARIETES
    worksheet_varietes.write(0, 0, 'ID')
    worksheet_varietes.write(0, 1, 'NOM')
    worksheet_varietes.write(0, 2, 'SLUG')

    produits = Variete.objects.all()
    row = 1
    for produit in produits:
        worksheet_varietes.write(row, 0, produit.id)
        worksheet_varietes.write(row, 1, produit.nom)
        worksheet_varietes.write(row, 2, produit.slug)
        row += 1

    # PORTEGREFFES
    worksheet_portegreffes.write(0, 0, 'ID')
    worksheet_portegreffes.write(0, 1, 'NOM')
    worksheet_portegreffes.write(0, 2, 'SLUG')

    produits = PorteGreffe.objects.all()
    row = 1
    for produit in produits:
        worksheet_portegreffes.write(row, 0, produit.id)
        worksheet_portegreffes.write(row, 1, produit.nom)
        worksheet_portegreffes.write(row, 2, produit.slug)
        row += 1

    # SPECS
    worksheet_specialites.write(0, 0, 'ID')
    worksheet_specialites.write(0, 1, 'NOM')
    worksheet_specialites.write(0, 2, 'SLUG')

    produits = Spec.objects.all()
    row = 1
    for produit in produits:
        worksheet_specialites.write(row, 0, produit.id)
        worksheet_specialites.write(row, 1, produit.nom)
        worksheet_specialites.write(row, 2, produit.slug)
        row += 1

    workbook.close()
    output.seek(0)

    filename = 'ExportProduits.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


@login_required
@staff_member_required
def export_produits_csv(request):
    if request.method == 'POST':
        categorie = request.POST.get('categorie')
        if categorie is None:
            message = "Catégorie manquante ... Veuillez réessayer !"
            messages.error(request, message)
            return redirect('onlineshop:export-produits-xls')

        if categorie == "PRODUITS":
            produit_resource = ProduitResource()
            filename = "Produits.csv"
        if categorie == "ESPECES":
            produit_resource = EspeceResource()
            filename = "Especes.csv"
        if categorie == "VARIETES":
            produit_resource = VarieteResource()
            filename = "Varietes.csv"
        if categorie == "PORTE-GREFFES":
            produit_resource = PorteGreffeResource()
            filename = "Portegreffes.csv"
        if categorie == "SPECIALITES":
            produit_resource = SpecResource()
            filename = "Specialites.csv"
        if categorie == "COULEURS":
            produit_resource = CouleurResource()
            filename = "Couleurs.csv"
        if categorie == "GREFFONS":
            produit_resource = GreffonResource()
            filename = "Greffons.csv"

        dataset = produit_resource.export()
        response = HttpResponse(dataset.csv, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename='+filename
        return response

    previous_page = reverse('onlineshop:onlineshop-administration')
    return render(request, 'onlineshop/export_produit.html', {'previous_page': previous_page})


@login_required
@staff_member_required
def import_produits_csv(request):
    previous_page = reverse('onlineshop:onlineshop-administration')
    if request.method == 'POST':
        categorie = request.POST.get('categorie')
        if categorie is None:
            message = "Catégorie manquante ... Veuillez réessayer !"
            messages.error(request, message)
            return redirect('onlineshop:import-produits-xls')

        if categorie == "PRODUITS":
            produit_resource = ProduitResource()
        if categorie == "ESPECES":
            produit_resource = EspeceResource()
        if categorie == "VARIETES":
            produit_resource = VarieteResource()
        if categorie == "PORTE-GREFFES":
            produit_resource = PorteGreffeResource()
        if categorie == "SPECIALITES":
            produit_resource = SpecResource()
        if categorie == "COULEURS":
            produit_resource = CouleurResource()
        if categorie == "GREFFONS":
            produit_resource = GreffonResource()

        dataset = Dataset()
        new_datas = request.FILES['myfile']
        imported_data = dataset.load(new_datas.read().decode(), format='csv')
        result = produit_resource.import_data(dataset, dry_run=True)  # Test the data import

        if not result.has_errors():
            produit_resource.import_data(dataset, dry_run=False)  # Actually import now
            message = "Fichier importé avec succès !"
            messages.success(request, message)
            return redirect('onlineshop:onlineshop-administration')

    return render(request, 'onlineshop/import_produit.html', {'previous_page': previous_page})


@login_required
@staff_member_required
def import_produits_csv_backup(request):
    previous_page = reverse('onlineshop:onlineshop-administration')
    if request.method == 'POST':
        produit_resource = ProduitResource()
        # greffon_resource = GreffonResource()

        dataset = Dataset()
        new_datas = request.FILES['myfile']
        categorie = request.POST.get('categorie')
        if categorie is None:
            message = "Catégorie manquante ... Veuillez réessayer !"
            messages.error(request, message)
            return redirect('onlineshop:import-produits-xls')

        if categorie == "PRODUITS":
            col_list = ["id", "stock", "stock_bis", "stock_future"]
            df = pd.read_csv(new_datas, usecols=col_list)
            print(df['id'], df['stock'])

        if categorie == "GREFFONS":
            col_list = ["id", "produit", "greffons", "comm", "objectif", "realise", "reussi"]
            df = pd.read_csv(new_datas, usecols=col_list)
            print(df['id'], df['greffons'])

        # imported_data_produit = dataset.load(new_datas.read().decode(), format='csv')
        # result = produit_resource.import_data(dataset, dry_run=True)  # Test the data import

        # if not result.has_errors():
        #     produit_resource.import_data(dataset, dry_run=False)  # Actually import now
        #     message = "Fichier importé avec succès !"
        #     messages.success(request, message)
        #     return redirect('onlineshop:onlineshop-administration')

    return render(request, 'onlineshop/import_produit.html', {'previous_page': previous_page})


@login_required
@staff_member_required
def export_greffons_xls(request):
    output = io.BytesIO()

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output)
    worksheet_greffons = workbook.add_worksheet("GREFFONS")
    worksheet_couleurs = workbook.add_worksheet("COULEURS")

    # GREFFONS
    worksheet_greffons.write(0, 0, 'ID')
    worksheet_greffons.write(0, 1, 'PRODUIT')
    worksheet_greffons.write(0, 2, 'GREFFONS')
    worksheet_greffons.write(0, 3, 'OBJECTIFS')
    worksheet_greffons.write(0, 4, 'COMM')
    worksheet_greffons.write(0, 5, 'REALISES')
    worksheet_greffons.write(0, 6, 'REUSSIS')
    worksheet_greffons.write(0, 7, 'DATE')
    worksheet_greffons.write(0, 8, 'COULEUR')
    worksheet_greffons.write(0, 9, 'RANG')
    worksheet_greffons.write(0, 10, 'PERIODE')

    greffons = Greffons.objects.all()
    row = 1
    for greffon in greffons:
        worksheet_greffons.write(row, 0, greffon.id)
        worksheet_greffons.write(row, 1, greffon.produit.id)
        worksheet_greffons.write(row, 2, greffon.greffons)
        worksheet_greffons.write(row, 3, greffon.objectif)
        worksheet_greffons.write(row, 4, greffon.comm)
        worksheet_greffons.write(row, 5, greffon.realise)
        worksheet_greffons.write(row, 6, greffon.reussi)
        worksheet_greffons.write(row, 7, greffon.date)
        worksheet_greffons.write(row, 8, greffon.couleur.id)
        worksheet_greffons.write(row, 9, greffon.rang)
        worksheet_greffons.write(row, 10, greffon.inventaire.id)
        row += 1

    # COULEURS
    worksheet_couleurs.write(0, 0, 'ID')
    worksheet_couleurs.write(0, 1, 'NOM')
    worksheet_couleurs.write(0, 2, 'COULEUR')

    couleurs = Couleur.objects.all()
    row = 1
    for couleur in couleurs:
        worksheet_couleurs.write(row, 0, couleur.id)
        worksheet_couleurs.write(row, 1, couleur.nom)
        worksheet_couleurs.write(row, 2, couleur.couleur)
        row += 1

    workbook.close()
    output.seek(0)

    filename = 'ExportGreffons.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


@login_required
@staff_member_required
def export_produits_xls_custom(request):
    output = io.BytesIO()

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output)
    worksheet_produits = workbook.add_worksheet("PRODUITS")

    # PRODUITS
    worksheet_produits.write(0, 0, 'ID')
    worksheet_produits.write(0, 1, 'NOM')
    worksheet_produits.write(0, 2, 'DESCRIPTION')
    worksheet_produits.write(0, 3, 'PRIX')
    worksheet_produits.write(0, 4, 'STOCK')
    worksheet_produits.write(0, 5, 'STOCK BIS')
    worksheet_produits.write(0, 6, 'AVAILABLE')
    worksheet_produits.write(0, 7, 'ESPECE')
    worksheet_produits.write(0, 8, 'VARIETE')
    worksheet_produits.write(0, 9, 'PORTEGREFFE')
    worksheet_produits.write(0, 10, 'SPEC')
    worksheet_produits.write(0, 11, 'GAF')

    produits = Produit.objects.all()

    row = 1
    for produit in produits:
        worksheet_produits.write(row, 0, produit.id)
        worksheet_produits.write(row, 1, produit.nom)
        worksheet_produits.write(row, 2, produit.description)
        worksheet_produits.write(row, 3, produit.prix)
        worksheet_produits.write(row, 4, produit.stock)
        worksheet_produits.write(row, 5, produit.stock_bis)
        worksheet_produits.write(row, 6, produit.available)
        worksheet_produits.write(row, 7, produit.espece.nom)
        worksheet_produits.write(row, 8, produit.variete.nom)
        worksheet_produits.write(row, 9, produit.portegreffe.nom)
        if not produit.spec is None:
            worksheet_produits.write(row, 10, produit.spec.nom)
        worksheet_produits.write(row, 11, produit.gaf)

        row += 1

    workbook.close()
    output.seek(0)

    filename = 'ExportProduitsCustom.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


# IMPORT DES PRODUITS A PARTIR D'UN FICHIER EXCEL
@login_required
@staff_member_required
def import_produits_xls(request):

    previous_page = reverse('onlineshop:onlineshop-administration')

    try:
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            categorie = request.POST.get('categorie')
            message = ""

            wb = load_workbook(filename=settings.CONTENT_DIR + excel_file, read_only=True)
            ws = wb[categorie]
            max_col = ws.max_column
            max_row = ws.max_row
            print(max_row)

            # REMOVE DATA FROM TABLE PRODUITS
            if request.POST.get('delete_data', True):
                if categorie == "PRODUITS":
                    Produit.objects.all().delete()
                if categorie == "ESPECES":
                    Espece.objects.all().delete()
                if categorie == "VARIETES":
                    Variete.objects.all().delete()
                if categorie == "PORTE-GREFFES":
                    PorteGreffe.objects.all().delete()
                if categorie == "SPECIALITES":
                    Spec.objects.all().delete()

            for i in range(2, max_row + 1):
                if categorie == "PRODUITS":
                    obj = Produit.objects.create(
                        id=ws.cell(row=i, column=1).value,
                        nom=ws.cell(row=i, column=2).value,
                        slug=ws.cell(row=i, column=3).value,
                        description=ws.cell(row=i, column=4).value,
                        prix=ws.cell(row=i, column=5).value,
                        stock=ws.cell(row=i, column=6).value,
                        stock_bis=ws.cell(row=i, column=7).value,
                        available=ws.cell(row=i, column=8).value,
                        espece=get_object_from_id(ws.cell(row=i, column=9).value, 'espece'),
                        variete=get_object_from_id(ws.cell(row=i, column=10).value, 'variete'),
                        portegreffe=get_object_from_id(ws.cell(row=i, column=11).value, 'portegreffe'),
                        spec=get_object_from_id(ws.cell(row=i, column=12).value, 'spec'),
                        gaf=ws.cell(row=i, column=13).value,
                    )

                if categorie == "ESPECES":
                    obj = Espece.objects.create(
                        id=ws.cell(row=i, column=1).value,
                        nom=ws.cell(row=i, column=2).value,
                        slug=ws.cell(row=i, column=3).value
                    )
                    # message = "Fichier d'espèces importé avec succès (" + str(max_row) + " espèces importées)"

                if categorie == "VARIETES":
                    obj = Variete.objects.create(
                        id=ws.cell(row=i, column=1).value,
                        nom=ws.cell(row=i, column=2).value,
                        slug=ws.cell(row=i, column=3).value
                    )
                    # message = "Fichier de variétés importé avec succès (" + str(max_row) + " varietés importées)"

                if categorie == "PORTE-GREFFES":
                    obj = PorteGreffe.objects.create(
                        id=ws.cell(row=i, column=1).value,
                        nom=ws.cell(row=i, column=2).value,
                        slug=ws.cell(row=i, column=3).value
                    )
                    # message = "Fichier de porte-greffe importé avec succès (" + str(max_row) + " porte-greffes importés)"

                if categorie == "SPECIALITES":
                    obj = Spec.objects.create(
                        id=ws.cell(row=i, column=1).value,
                        nom=ws.cell(row=i, column=2).value,
                        slug=ws.cell(row=i, column=3).value
                    )
                    # message = "Fichier de spécialités importé avec succès (" + str(max_row) + " spécialités importées)"

                obj.save()
            print(max_row, "produits importés !")
            message = "Fichier importé avec succès (" + str(max_row) + " données importées)"
            messages.success(request, message)
            wb.close()

            return render(request, 'onlineshop/import_produit.html', {'previous_page': previous_page})

    except Exception as identifier:
        message = format_html("Une erreur est survenue ... merci de réessayer !<br>[ <i>" + str(identifier) + "</i> ]")
        messages.error(request, message)
        print(identifier)

    return render(request, 'onlineshop/import_produit.html', {'previous_page': previous_page})
    # return response


# IMPORT DES PRODUITS A PARTIR D'UN FICHIER EXCEL
@login_required
@staff_member_required
def import_produits_xls_pandas(request):

    previous_page = reverse('onlineshop:onlineshop-administration')

    try:
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            categorie = request.POST.get('categorie')
            message = ""
            empexceldata = pd.read_excel(settings.CONTENT_DIR + excel_file, sheet_name=categorie)
            dbframe = empexceldata

            # REMOVE DATA FROM TABLE PRODUITS
            if request.POST.get('delete_data', True):
                if categorie == "PRODUITS":
                    pass
                    # Produit.objects.all().delete()
                if categorie == "ESPECES":
                    Espece.objects.all().delete()
                if categorie == "VARIETES":
                    Variete.objects.all().delete()
                if categorie == "PORTE-GREFFES":
                    PorteGreffe.objects.all().delete()
                if categorie == "SPECIALITES":
                    Spec.objects.all().delete()
            max_row = 0
            for dbframe in dbframe.itertuples():
                if categorie == "PRODUITS":
                    obj = Produit.objects.create(
                        id=dbframe.id,
                        nom=dbframe.nom,
                        slug=dbframe.slug,
                        description=dbframe.description,
                        prix=dbframe.prix,
                        stock=dbframe.stock,
                        stock_bis=dbframe.stock_bis,
                        available=dbframe.available,
                        espece=get_object_from_id(dbframe.espece, 'espece'),
                        variete=get_object_from_id(dbframe.variete, 'variete'),
                        portegreffe=get_object_from_id(dbframe.portegreffe, 'portegreffe'),
                        spec=get_object_from_id(dbframe.spec, 'spec'),
                        gaf=dbframe.gaf,
                    )
                # obj.save()
                max_row += 1
            print(max_row, "produits importés !")
            message = "Fichier importé avec succès (" + str(max_row) + " données importées)"
            messages.success(request, message)

            return render(request, 'onlineshop/import_produit.html', {'previous_page': previous_page})

    except Exception as identifier:
        message = format_html("Une erreur est survenue ... merci de réessayer !<br>[ <i>" + str(identifier) + "</i> ]")
        messages.error(request, message)
        print(identifier)

    return render(request, 'onlineshop/import_produit.html', {'previous_page': previous_page})
    # return response


# IMPORT DES GREFFONS A PARTIR D'UN FICHIER EXCEL
@login_required
@staff_member_required
def import_greffons_xls(request):

    previous_page = reverse('onlineshop:onlineshop-administration')

    try:
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            fs = FileSystemStorage()
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = fs.url(filename)
            excel_file = uploaded_file_url
            message = format_html("Fichier de greffons importé avec succès<br/>")

            wb = load_workbook(filename=settings.CONTENT_DIR + excel_file, read_only=True)
            ws_greffons = wb["GREFFONS"]
            ws_couleurs = wb["COULEURS"]
            max_col_greffons = ws_greffons.max_column
            max_row_greffons = ws_greffons.max_row
            max_col_couleurs = ws_couleurs.max_column
            max_row_couleurs = ws_couleurs.max_row

            # REMOVE DATA FROM TABLE PRODUITS
            if request.POST.get('delete_data', True):
                    Greffons.objects.all().delete()
                    Couleur.objects.all().delete()

            # GREFFONS
            for i in range(2, ws_greffons + 1):
                obj = Produit.objects.create(
                    id=ws_greffons.cell(row=i, column=1).value,
                    produit=get_object_from_id(ws_greffons.cell(row=i, column=2).value, 'produit'),
                    greffons=ws_greffons.cell(row=i, column=3).value,
                    comm=ws_greffons.cell(row=i, column=4).value,
                    objectif=ws_greffons.cell(row=i, column=5).value,
                    realise=ws_greffons.cell(row=i, column=6).value,
                    reussi=ws_greffons.cell(row=i, column=7).value,
                    date=ws_greffons.cell(row=i, column=8).value,
                    couleur=get_object_from_id(ws_greffons.cell(row=i, column=9).value, 'couleur'),
                    rang=ws_greffons.cell(row=i, column=10).value,
                    periode=get_object_from_id(ws_greffons.cell(row=i, column=11).value, 'inventaire'),
                )
                message += format_html("- " + str(max_row_greffons) + " greffons importés<br/>")

                # GREFFONS
                for i in range(2, ws_couleurs + 1):
                    obj = Couleur.objects.create(
                        id=ws_couleurs.cell(row=i, column=1).value,
                        nom=ws_couleurs.cell(row=i, column=2).value,
                        couleur=ws_couleurs.cell(row=i, column=3).value,
                    )
                message += format_html("- " + str(max_row_couleurs) + " couleurs importés")

                obj.save()

            messages.success(request, message)
            wb.close()

            return render(request, 'onlineshop/import_greffons.html', {'previous_page': previous_page})

    except Exception as identifier:
        message = format_html("Une erreur est survenue ... merci de réessayer !<br>[ <i>" + str(identifier) + "</i> ]")
        messages.error(request, message)
        print(identifier)

    return render(request, 'onlineshop/import_greffons.html', {'previous_page': previous_page})
    # return response


@login_required
@staff_member_required
def comm_greffons(request):
    if request.method == 'POST':
        qte = get_qte_pre_commande(request.POST.get('produit'))
        return HttpResponse(qte)


@login_required
@staff_member_required
def reset_stock(request):
    cancel_statut = Statut.objects.get(nom='Annulée')
    done_statut = Statut.objects.get(nom='Terminée')
    inprogress_statut = Statut.objects.get(nom='En cours')
    future_statut = Statut.objects.get(nom='Pré-commande')
    form_statut = Statut.objects.filter(nom__in=['Validée', 'En cours'])
    commandes = Commande.objects.exclude(statut__in=[cancel_statut, done_statut, future_statut])
    context = {}

    if commandes.count() > 0:
        message = format_html("Impossible de réinitialiser les Stocks !<br>Certaines commandes ne sont pas terminées/annulées")
        messages.error(request, message)

        form = SearchOrderForm(
            initial={
                'statut': form_statut, })
        context = {
            'form': form,
            'formAction': 'order:manage-order',
            'previous_page': reverse('onlineshop:onlineshop-administration'),
        }
        return render(request, 'order/manage_order.html', context)
        # return redirect('order:manage-order')

    produits = Produit.objects.all()
    for produit in produits:

        old_final = produit.stock
        old_bis = produit.stock_bis
        old_future = produit.stock_future

        produit.stock = 0
        produit.stock_bis = 0
        # produit.stock_future = 0
        produit.save()

        if old_final > 0:
            log_produit(str(request.user), produit.pk, None, 'Edit', 'sf', old_final, 0)
        if old_bis != 0:
            log_produit(str(request.user), produit.pk, None, 'Edit', 'sb', old_bis, 0)

    message = format_html("Les stocks (initiaux, virtuels) ont bien été réinitalisés !")
    messages.success(request, message)

    return redirect('onlineshop:onlineshop-administration')


# MISE A JOUR DES STOCKS D'UN PRODUIT
@login_required
@staff_member_required
def edit_stock_produit(request):
    if request.method == 'POST' and request.is_ajax:
        data_string = request.POST.get('json_data')
        data_dict = json.loads(data_string)
        print(data_dict)

        produit_id = data_dict['produit_id']
        stock = data_dict['stock']
        stock_bis = data_dict['stock_bis']
        stock_future = data_dict['stock_future']

        try:
            produit = Produit.objects.get(id=produit_id)
        except Exception as e:
            print(e)
            messages.error(request, "Le produit n'existe pas !")
            return redirect('onlineshop:manage-produit')

        old_final = produit.stock
        old_bis = produit.stock_bis
        old_future = produit.stock_future

        if check_stock_value(stock) is not None:
            produit.stock = int(stock)
            if old_final != int(stock):
                log_produit(str(request.user), produit.pk, None, 'Edit', 'sf', old_final, int(stock))
        if check_stock_value(stock_bis) is not None:
            produit.stock_bis = int(stock_bis)
            if old_bis != int(stock_bis):
                log_produit(str(request.user), produit.pk, None, 'Edit', 'sb', old_bis, int(stock_bis))
        if check_stock_value(stock_future) is not None:
            produit.stock_future = int(stock_future)
            if old_future != int(stock_future):
                log_produit(str(request.user), produit.pk, None, 'Edit', 'sp', old_future, int(stock_future))
        if int(stock_bis) > int(stock):
            produit.stock = stock_bis
            if old_final != int(stock_bis):
                log_produit(str(request.user), produit.pk, None, 'Edit', 'sf', old_final, int(stock_bis))
        produit.save()

        messages.success(request, "Quantité mise à jour pour le produit ;)")
    return redirect('onlineshop:manage-produit')


@login_required
@staff_member_required
def warning_produit(request):
    pass
    form = SearchProduitForm

    if request.user.is_staff:
        inventaire = Inventaire.objects.filter(actif=True).first()
        produits_warning = get_list_produits_anomalie(inventaire)
        print("Produit en anomalie :", len(produits_warning))
        title = "Produits"
        header = "Liste des produits en anomalie"
        javascript = ""
        formAction = "onlineshop:manage-produit"
        previous_page = reverse('onlineshop:onlineshop-administration')

        paginator = Paginator(produits_warning, 50)
        page = request.GET.get('page', 1)
        try:
            produits_warning = paginator.page(page)
        except PageNotAnInteger:
            produits_warning = paginator.page(1)
        except EmptyPage:
            produits_warning = paginator.page(paginator.num_pages)

        context_header = {
            'header': header,
            'javascript': javascript,
        }
        context = {
            'previous_page': previous_page,
            'title': title,
            'context_header': context_header,
            'formAction': formAction,
            'form': form,
            'produits_warning': produits_warning,
        }
        return render(request, "onlineshop/manage_warning_produit.html", context)

    else:
        messages.error(request, "Vous n'avez pas les droits !")
    return redirect('onlineshop:produit-list')